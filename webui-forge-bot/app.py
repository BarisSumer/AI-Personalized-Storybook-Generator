# app.py
import os, io, csv, json, uuid, time, base64, threading, datetime as dt, re, sys, subprocess
from typing import List, Dict, Any, Optional
from flask import (
    Flask, request, redirect, url_for, flash, render_template_string, abort,
    Response, stream_with_context, send_file
)
import requests
from PIL import Image

APP_TITLE = "Kitap Yönetimi"
DATA_DIR = os.path.join(os.path.dirname(__file__), "data")
BOOKS_DIR = os.path.join(DATA_DIR, "books")
LOGS_DIR = os.path.join(DATA_DIR, "logs")
os.makedirs(BOOKS_DIR, exist_ok=True)
os.makedirs(LOGS_DIR, exist_ok=True)

DEFAULT_FACES_DIR = r"C:\faces"
DEFAULT_POSES_DIR = r"C:\poses"
DEFAULT_OUT_DIR = r"C:\out"
os.makedirs(DEFAULT_OUT_DIR, exist_ok=True)

SD_BASE = os.environ.get("SD_BASE", "http://127.0.0.1:7861")

# UI runner betiğinin yolu (gerekirse değiştir)
RUNNER_PATH = os.path.join(os.path.dirname(__file__), "runner_ui_prompts.py")
RUNNER_HEADLESS = os.environ.get("FORGE_UI_HEADLESS", "0") == "1"

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "dev-secret")

from jinja2 import DictLoader, ChoiceLoader

BASE_HTML = r"""<!doctype html><html lang="tr"><head><meta charset="utf-8">
<title>{{ title or 'Uygulama' }}</title><meta name="viewport" content="width=device-width, initial-scale=1">
<style>
:root { --bg:#0f1216; --panel:#151a21; --ink:#e8e8ea; --muted:#9aa3af; --ok:#22c55e; --warn:#f59e0b; --err:#ef4444; --acc:#3b82f6; }
* { box-sizing: border-box; } body { margin:0; font-family: ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial; background:var(--bg); color:var(--ink); }
a { color: var(--acc); text-decoration: none; }
header { padding: 16px 20px; background: #0b0e12; border-bottom: 1px solid #1f2732; position: sticky; top:0; z-index:10;}
header h1 { margin:0; font-size: 18px; } nav { display:flex; gap:12px; margin-top:8px; }
.container { max-width: 1140px; margin: 24px auto; padding: 0 16px; }
.panel { background: var(--panel); border:1px solid #1f2732; border-radius: 12px; padding:16px; }
.row { display:flex; gap:12px; flex-wrap: wrap; } .col { flex:1 1 260px; min-width: 260px; }
label { display:block; font-size:12px; color:var(--muted); margin-bottom:6px; }
input[type=text], input[type=number], textarea, select { width:100%; background:#10151b; border:1px solid #1f2732; color:var(--ink); border-radius:10px; padding:10px 12px; outline:none; }
textarea { min-height: 90px; resize: vertical; }
.switch { display:flex; align-items:center; gap:8px; }
.btnrow { display:flex; gap:10px; flex-wrap:wrap; margin-top:12px; }
.btn { background:#0f1720; border:1px solid #273141; color:var(--ink); padding:10px 14px; border-radius:10px; cursor:pointer; }
.btn.primary { background:#1a2a43; border-color:#2b4161; } .btn.danger{ background:#2a1518; border-color:#4b1d24; color:#fca5a5; } .btn.ok{ background:#142a1c; border-color:#1f4d2e; color:#86efac; }
table { width:100%; border-collapse: collapse; } th, td { text-align:left; padding:10px; border-bottom:1px solid #1f2732; }
.flash { margin: 12px 0; padding: 10px 12px; border-radius:10px; border:1px solid #273141; background:#0f1720; }
.muted { color: var(--muted); font-size: 12px; } .status { padding: 2px 8px; border-radius: 999px; border:1px solid #273141; font-size: 12px; } .hl { color:#fbbf24; }
</style></head><body>
<header><h1>{{ title or 'Uygulama' }}</h1>
<nav><a href="{{ url_for('ui_list_books') }}">📚 Kitaplar</a><a href="{{ url_for('ui_new_book') }}">➕ Kitap Ekle</a></nav></header>
<div class="container">{% with messages = get_flashed_messages() %}{% if messages %}{% for m in messages %}<div class="flash">{{ m }}</div>{% endfor %}{% endif %}{% endwith %}{% block content %}{% endblock %}</div>
</body></html>"""
app.jinja_loader = ChoiceLoader([DictLoader({"base.html": BASE_HTML}), app.jinja_loader])


def now_iso(): return dt.datetime.now().isoformat(timespec="seconds")
def book_file(book_id: str): return os.path.join(BOOKS_DIR, f"{book_id}.json")
def write_book(book: dict):
    with open(book_file(book["id"]), "w", encoding="utf-8") as f:
        json.dump(book, f, ensure_ascii=False, indent=2)

def ensure_settings_defaults(book: dict, save=False):
    if not book: return book
    s = book.setdefault("settings", {})
    changed = False
    def setdef(k, v):
        nonlocal changed
        if k not in s:
            s[k] = v; changed = True
    setdef("data_source", "excel")
    setdef("output_root", DEFAULT_OUT_DIR)
    setdef("excel_path", "")
    setdef("col_photo", "@photo")
    setdef("col_first", "student name")
    setdef("col_last", "student last name")
    setdef("col_class", "class")
    setdef("faces_dir", DEFAULT_FACES_DIR)
    setdef("poses_dir", DEFAULT_POSES_DIR)  # kitap seviyesinde Yedek/fallback
    setdef("col_out", "out")
    if save and changed and book.get("id"):
        write_book(book)
    return book

def read_book(book_id: str):
    p = book_file(book_id)
    if not os.path.exists(p): return None
    with open(p, "r", encoding="utf-8") as f:
        data = json.load(f)
    return ensure_settings_defaults(data, save=True)

def list_books():
    out = []
    for n in os.listdir(BOOKS_DIR):
        if n.endswith(".json"):
            try:
                with open(os.path.join(BOOKS_DIR, n), "r", encoding="utf-8") as f:
                    out.append(ensure_settings_defaults(json.load(f), save=True))
            except: pass
    out.sort(key=lambda b: b.get("updated_at", b.get("created_at", "")), reverse=True)
    return out

def api_get(url_path: str) -> Any:
    try:
        r = requests.get(SD_BASE + url_path, timeout=10)
        r.raise_for_status()
        return r.json()
    except Exception:
        return None

def api_models() -> List[str]:
    js = api_get("/sdapi/v1/sd-models") or []
    names = [m.get("model_name") or m.get("title") or "" for m in js]
    return [n for n in names if n]

def api_samplers() -> List[str]:
    js = api_get("/sdapi/v1/samplers") or []
    return [m.get("name") for m in js if m.get("name")]

def api_styles() -> List[str]:
    js = api_get("/sdapi/v1/prompt-styles") or []
    return [s.get("name") for s in js if s.get("name")]

def api_cn_model_list() -> List[str]:
    js = api_get("/controlnet/model_list") or {}
    return js.get("model_list", []) if isinstance(js, dict) else []

def api_cn_module_list() -> List[str]:
    js = api_get("/controlnet/module_list") or {}
    return js.get("module_list", []) if isinstance(js, dict) else []

def _ensure_data_uri(b64_plain_or_data_uri: str) -> str:
    s = (b64_plain_or_data_uri or "").strip()
    if s.startswith("data:image"):
        return s
    return "data:image/png;base64," + s

def _pil_to_data_uri(img: Image.Image) -> str:
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return "data:image/png;base64," + base64.b64encode(buf.getvalue()).decode("utf-8")

def reactor_available() -> bool:
    try:
        r = requests.get(SD_BASE + "/reactor/models", timeout=5)
        return r.ok
    except Exception:
        return False

def reactor_swap(face_b64_plain: str, target_pil: Image.Image, opts: Dict[str, Any]) -> Image.Image:
    body = {
        "source_image": _ensure_data_uri(face_b64_plain),
        "target_image": _pil_to_data_uri(target_pil),
        # Indeksler tek öğeli liste istiyor
        "source_faces_index": [int(opts.get("source_face_index", 0))],
        "face_index": [int(opts.get("face_index", 0))],
        # Varsayılanlar (UI’ye en yakın sade ayarlar)
        "upscaler": opts.get("upscaler", "None"),
        "scale": int(opts.get("scale", 1)),
        "upscale_visibility": float(opts.get("upscale_visibility", 1)),
        "face_restorer": opts.get("face_restorer", "None"),
        "restorer_visibility": float(opts.get("restorer_visibility", 1)),
        "restore_first": int(opts.get("restore_first", 0)),
        "model": opts.get("model", "inswapper_128.onnx"),
        "gender_source": int(opts.get("gender_source", 0)),
        "gender_target": int(opts.get("gender_target", 0)),
        "save_to_file": 0,
        "result_file_path": ""
    }
    r = requests.post(SD_BASE + "/reactor/image", json=body, timeout=120)
    r.raise_for_status()
    js = r.json() or {}
    img_b64 = js.get("image") or ""
    # Bazı sürümler data-uri değil, düz b64 döndürür
    if img_b64.startswith("data:image"):
        img_b64 = img_b64.split(",", 1)[-1]
    return Image.open(io.BytesIO(base64.b64decode(img_b64))).copy()



BOOK_LIST_HTML = r"""
{% extends "base.html" %}{% block content %}
  <div class="panel">
    <div class="btnrow"><a class="btn primary" href="{{ url_for('ui_new_book') }}">➕ Yeni Kitap</a></div>
    <table><thead><tr><th>Ad</th><th>Sayfa</th><th>Kaynak</th><th>Çıkış</th><th>Güncellendi</th><th style="width:420px"></th></tr></thead>
      <tbody>{% for b in books %}
        <tr>
          <td><a href="{{ url_for('ui_book_pages', book_id=b.id) }}">{{ b.name }}</a></td>
          <td>{{ (b.pages or [])|length }}</td>
          <td class="muted">{{ b.settings.data_source }}{% if b.settings.data_source=='excel' %} · {{ b.settings.excel_path }}{% else %} · {{ b.settings.faces_dir }}{% endif %}</td>
          <td class="muted">{{ b.settings.output_root }}</td>
          <td class="muted">{{ b.updated_at or b.created_at }}</td>
          <td>
            <div class="btnrow">
              <a class="btn" href="{{ url_for('ui_book_pages', book_id=b.id) }}">📄 Sayfalar</a>
              <form method="post" action="{{ url_for('ui_run_book', book_id=b.id) }}" style="display:inline;"><button class="btn ok" type="submit">▶️ API'den Çalıştır</button></form>
              <form method="post" action="{{ url_for('ui_run_book_ui', book_id=b.id) }}" style="display:inline;"><button class="btn" type="submit">🖥️ Forge Arayüzünden Çalıştır</button></form>
              <form method="post" action="{{ url_for('ui_delete_book', book_id=b.id) }}" style="display:inline;" onsubmit="return confirm('Kitap silinsin mi?')"><button class="btn danger" type="submit">🗑 Sil</button></form>
            </div>
          </td>
        </tr>
      {% else %}<tr><td colspan="6" class="muted">Henüz kitap yok.</td></tr>{% endfor %}</tbody>
    </table>
  </div>
{% endblock %}
"""

BOOK_FORM_HTML = r"""
{% extends "base.html" %}{% block content %}
  <form method="post"><div class="panel">
    <div class="row"><div class="col"><label>Kitap Adı</label><input type="text" name="name" value="{{ b.name or '' }}" required></div></div>
    <h3>Veri Kaynağı</h3>
    <div class="row">
      <div class="col"><label>Kaynak tipi</label>
        <select name="data_source">
          <option value="excel"  {% if b.settings.data_source=='excel' %}selected{% endif %}>Excel/CSV</option>
          <option value="folders"{% if b.settings.data_source=='folders' %}selected{% endif %}>Klasör (yüz dosyaları)</option>
        </select>
      </div>
      <div class="col"><label>Çıkış kök klasörü</label><input type="text" name="output_root" value="{{ b.settings.output_root or '' }}" placeholder="C:\Users\...\Desktop\hikayeler"></div>
    </div>

    <h3>Excel/CSV ayarları</h3>
    <div class="row">
      <div class="col"><label>Excel dosyası yolu</label><input type="text" name="excel_path" value="{{ b.settings.excel_path or '' }}" placeholder="C:\...\students.xlsx"></div>
      <div class="col"><label>Foto sütunu (varsayılan: @photo)</label><input type="text" name="col_photo" value="{{ b.settings.col_photo or '@photo' }}"></div>
    </div>
    <div class="row">
      <div class="col"><label>Ad sütunu</label><input type="text" name="col_first" value="{{ b.settings.col_first or '' }}"></div>
      <div class="col"><label>Soyad sütunu</label><input type="text" name="col_last" value="{{ b.settings.col_last or '' }}"></div>
      <div class="col"><label>Sınıf sütunu</label><input type="text" name="col_class" value="{{ b.settings.col_class or '' }}"></div>
    </div>

    <h3>Klasör yöntemi (opsiyonel)</h3>
    <div class="row">
      <div class="col"><label>Çocuk yüzleri klasörü (faces_dir)</label><input type="text" name="faces_dir" value="{{ b.settings.faces_dir or '' }}" placeholder="C:\faces"><div class="muted">Excel'de foto yolu göreliyse bununla birleşir.</div></div>
      <div class="col"><label>Poz klasörü veya dosyası (varsayılan)</label><input type="text" name="poses_dir" value="{{ b.settings.poses_dir or '' }}" placeholder="C:\poses\pose01 veya C:\poses\pose01\pose.png"><div class="muted">Sayfa bazında boş bırakılırsa bu değer kullanılır.</div></div>
    </div>

    <div class="btnrow"><button class="btn ok" type="submit">💾 Kaydet</button><a class="btn" href="{{ url_for('ui_list_books') }}">İptal</a>{% if b.id %}<a class="btn" href="{{ url_for('ui_book_pages', book_id=b.id) }}">📄 Sayfalar</a>{% endif %}</div>
  </div></form>
{% endblock %}
"""

PAGES_HTML = r"""
{% extends "base.html" %}{% block content %}
  <div class="panel">
    <div class="row">
      <div class="col"><h2 style="margin:0">{{ b.name }}</h2></div>
      <div class="col" style="text-align:right"><a class="btn" href="{{ url_for('ui_edit_book', book_id=b.id) }}">⚙️ Kitap Ayarları</a><a class="btn" href="{{ url_for('ui_list_books') }}">↩︎ Geri</a></div>
    </div>
    <p class="muted">Kitap ayarlarında <span class="hl">veri kaynağı ve klasörler</span> var. Tüm SD ayarları sayfa bazında.</p>

    <div class="btnrow" style="margin-top:6px">
      <form method="post" action="{{ url_for('ui_run_book', book_id=b.id) }}"><button class="btn ok" type="submit">▶️ API'den Çalıştır</button></form>
      <form method="post" action="{{ url_for('ui_run_book_ui', book_id=b.id) }}"><button class="btn" type="submit">🖥️ Forge Arayüzünden Çalıştır</button></form>
      {% if last_job_id %}<a class="btn" href="{{ url_for('ui_job_status', job_id=last_job_id) }}">📝 Son İş: {{ last_job_id[:8] }} <span class="status">{{ last_job_status }}</span></a>{% endif %}
    </div>

    <h3>Sayfa Ekle</h3>
    <form method="post" action="{{ url_for('ui_add_page', book_id=b.id) }}">
      <div class="row">
        <div class="col"><label>Sayfa No</label><input type="number" name="index" min="1" value="{{ (b.pages|length) + 1 }}"></div>
        <div class="col"><label>Seed</label><input type="number" name="seed" value="-1"><div class="muted">-1 = rastgele</div></div>
        <div class="col"><label>ControlNet Kullan</label><div class="switch"><input type="checkbox" name="use_controlnet" checked></div></div>
      </div>

      <div class="row">
        <div class="col">
          <label>Styles (çoklu seç)</label>
          <select name="styles" multiple size="8" style="height:auto;">
            {% for st in styles_list %}<option value="{{ st }}">{{ st }}</option>{% endfor %}
          </select>
          <div class="muted">Birden fazla seçim için Ctrl / ⌘ kullanın.</div>
        </div>
      </div>

      <div class="row">
        <div class="col"><label>Checkpoint</label><select name="checkpoint">{% for m in models %}<option value="{{m}}">{{m}}</option>{% endfor %}</select></div>
        <div class="col"><label>Sampling Method</label><select name="sampling_method">{% for s in samplers %}<option value="{{s}}" {% if s=='Euler a' %}selected{% endif %}>{{s}}</option>{% endfor %}</select></div>
        <div class="col"><label>Sampling Steps</label><input type="number" name="sampling_steps" min="1" max="200" step="1" value="12"></div>
      </div>

      <div class="row">
        <div class="col"><label>Width</label><input type="number" name="width" min="64" max="2048" step="1" value="1980"></div>
        <div class="col"><label>Height</label><input type="number" name="height" min="64" max="2048" step="1" value="1020"></div>
        <div class="col"><label>CFG Scale</label><input type="number" name="cfg_scale" min="1" max="30" step="0.5" value="4.0"></div>
      </div>

      <div class="row"><div class="col" style="flex:1 1 100%"><label>Pozitif Prompt</label><textarea name="prompt" required></textarea></div></div>
      <div class="row"><div class="col" style="flex:1 1 100%"><label>Negatif Prompt</label><textarea name="negative_prompt"></textarea></div></div>

      <h4>ControlNet (Instant-ID)</h4>
      <div class="row">
        <div class="col">
          <label>Unit 0 Module</label>
          <select name="cn0_module">
            {% for m in cn_modules %}<option value="{{m}}" {% if m=='InsightFace (InstantID)' %}selected{% endif %}>{{m}}</option>{% endfor %}
          </select>
        </div>
        <div class="col">
          <label>Unit 0 Model</label>
          <select name="cn0_model">
            {% for m in cn_models %}<option value="{{m}}" {% if 'ip-adapter_instant_id_sdxl' in m %}selected{% endif %}>{{m}}</option>{% endfor %}
          </select>
        </div>
        <div class="col">
          <label>Unit 0 Resize mode</label>
          <select name="cn0_resize">
            <option value="0">Just Resize</option>
            <option value="1" selected>Crop and Resize</option>
            <option value="2">Resize and Fill</option>
          </select>
        </div>
      </div>

      <div class="row">
        <div class="col">
          <label>Unit 0 Control Weight</label>
          <input type="number" name="cn0_weight" min="0" max="2" step="0.05" value="0.5">
        </div>
        <div class="col">
          <label>Unit 0 Control Mode</label>
          <select name="cn0_mode">
            <option value="0" selected>Balanced</option>
            <option value="1">My prompt is more important</option>
            <option value="2">ControlNet is more important</option>
          </select>
        </div>
      </div>

      <div class="row">
        <div class="col">
          <label>Unit 1 Module</label>
          <select name="cn1_module">
            {% for m in cn_modules %}<option value="{{m}}" {% if m=='instant_id_face_keypoints' %}selected{% endif %}>{{m}}</option>{% endfor %}
          </select>
        </div>
        <div class="col">
          <label>Unit 1 Model</label>
          <select name="cn1_model">
            {% for m in cn_models %}<option value="{{m}}" {% if 'control_instant_id_sdxl' in m %}selected{% endif %}>{{m}}</option>{% endfor %}
          </select>
        </div>
        <div class="col">
          <label>Unit 1 Resize mode</label>
          <select name="cn1_resize">
            <option value="0">Just Resize</option>
            <option value="1">Crop and Resize</option>
            <option value="2" selected>Resize and Fill</option>
          </select>
        </div>
      </div>

      <div class="row">
        <div class="col">
          <label>Unit 1 Control Weight</label>
          <input type="number" name="cn1_weight" min="0" max="2" step="0.05" value="0.5">
        </div>
        <div class="col">
          <label>Unit 1 Control Mode</label>
          <select name="cn1_mode">
            <option value="0" selected>Balanced</option>
            <option value="1">My prompt is more important</option>
            <option value="2">ControlNet is more important</option>
          </select>
        </div>
      </div>

      <h4>Poz (Sayfaya özel)</h4>
      <div class="row">
        <div class="col" style="flex:1 1 100%">
          <label>Poz (dosya veya klasör yolu)</label>
          <input type="text" name="pose_path" placeholder="C:\poses\pose03 veya C:\poses\pose03\pose.png">
          <div class="muted">Boş bırakılırsa kitap ayarlarında varsayılan <code>poses_dir</code> kullanılır.</div>
        </div>
      </div>

      <h4>REActor (FaceSwap)</h4>
      <div class="row">
        <div class="col">
          <label>REActor Kullan</label>
          <div class="switch"><input type="checkbox" name="use_reactor"></div>
          <div class="muted">Açıkken üretim sonrası swap uygulanır (Forge REActor). Aşağıda istersen basit ayarları JSON ver.</div>
        </div>
      </div>
      <div class="row">
        <div class="col" style="flex:1 1 100%">
          <label>REActor Ayarları (JSON, opsiyonel)</label>
          <textarea name="reactor_json" placeholder='{"model":"inswapper_128.onnx","face_index":0}'></textarea>
          <div class="muted">Desteklenen anahtarlar: <code>model</code>, <code>face_index</code>, <code>source_face_index</code>.</div>
        </div>
      </div>

      <div class="btnrow"><button class="btn ok" type="submit">➕ Ekle</button></div>
    </form>

    <h3 style="margin-top:24px">Sayfalar</h3>
    <table><thead><tr><th>#</th><th>Seed</th><th>W×H</th><th>Steps</th><th>Sampler</th><th>Checkpoint</th><th>CN</th><th>Poz</th><th>Pozitif Prompt</th><th></th></tr></thead>
      <tbody>
        {% for p in (b.pages or []) %}
          <tr>
            <td>{{ p.index }}</td><td>{{ p.seed }}</td><td>{{ p.width }}×{{ p.height }}</td><td>{{ p.sampling_steps }}</td>
            <td class="muted">{{ p.sampling_method }}</td><td class="muted">{{ p.checkpoint }}</td>
            <td>{{ 'Açık' if p.use_controlnet else 'Kapalı' }}</td>
            <td class="muted" title="{{ p.pose_path or '' }}">{{ (p.pose_path or '').split('\\')[-1].split('/')[-1] }}</td>
            <td class="muted" title="{{ p.prompt }}">{{ p.prompt[:80] ~ ('…' if p.prompt|length > 80 else '') }}</td>
            <td>
              <div class="btnrow">
                <a class="btn" href="{{ url_for('ui_edit_page', book_id=b.id, page_id=p.id) }}">✏️ Düzenle</a>
                <form method="post" action="{{ url_for('ui_delete_page', book_id=b.id, page_id=p.id) }}" onsubmit="return confirm('Sayfa silinsin mi?')"><button class="btn danger" type="submit">🗑 Sil</button></form>
              </div>
            </td>
          </tr>
        {% else %}<tr><td colspan="10" class="muted">Henüz sayfa yok.</td></tr>{% endfor %}
      </tbody>
    </table>
  </div>
{% endblock %}
"""


PAGE_EDIT_HTML = r"""
{% extends "base.html" %}{% block content %}
  <form method="post"><div class="panel">
    <h3>Sayfa Düzenle (#{{ p.index }})</h3>
    <div class="row">
      <div class="col"><label>Sayfa No</label><input type="number" name="index" value="{{ p.index }}"></div>
      <div class="col"><label>Seed</label><input type="number" name="seed" value="{{ p.seed }}"></div>
      <div class="col"><label>ControlNet Kullan</label><div class="switch"><input type="checkbox" name="use_controlnet" {% if p.use_controlnet %}checked{% endif %}></div></div>
    </div>

    <div class="row">
      <div class="col">
        <label>Styles (çoklu seç)</label>
        <select name="styles" multiple size="8" style="height:auto;">
          {% for st in styles_list %}
            <option value="{{ st }}" {% if p.styles and (st in p.styles) %}selected{% endif %}>{{ st }}</option>
          {% endfor %}
        </select>
        <div class="muted">Birden fazla seçim için Ctrl / ⌘ kullanın.</div>
      </div>
    </div>

    <div class="row">
      <div class="col"><label>Checkpoint</label><select name="checkpoint">{% for m in models %}<option value="{{m}}" {% if p.checkpoint==m %}selected{% endif %}>{{m}}</option>{% endfor %}</select></div>
      <div class="col"><label>Sampling Method</label><select name="sampling_method">{% for s in samplers %}<option value="{{s}}" {% if p.sampling_method==s %}selected{% endif %}>{{s}}</option>{% endfor %}</select></div>
      <div class="col"><label>Sampling Steps</label><input type="number" name="sampling_steps" min="1" max="200" step="1" value="{{ p.sampling_steps }}"></div>
    </div>
    <div class="row">
      <div class="col"><label>Width</label><input type="number" name="width" min="64" max="2048" step="1" value="{{ p.width }}"></div>
      <div class="col"><label>Height</label><input type="number" name="height" min="64" max="2048" step="1" value="{{ p.height }}"></div>
      <div class="col"><label>CFG Scale</label><input type="number" name="cfg_scale" min="1" max="30" step="0.5" value="{{ p.cfg_scale }}"></div>
    </div>
    <div class="row"><div class="col" style="flex:1 1 100%"><label>Pozitif Prompt</label><textarea name="prompt">{{ p.prompt }}</textarea></div></div>
    <div class="row"><div class="col" style="flex:1 1 100%"><label>Negatif Prompt</label><textarea name="negative_prompt">{{ p.negative_prompt }}</textarea></div></div>

        <h4>ControlNet (Instant-ID)</h4>
    <div class="row">
      <div class="col">
        <label>Unit 0 Module</label>
        <select name="cn0_module">
          {% for m in cn_modules %}
            <option value="{{m}}" {% if p.cn0_module==m %}selected{% endif %}>{{m}}</option>
          {% endfor %}
        </select>
      </div>
      <div class="col">
        <label>Unit 0 Model</label>
        <select name="cn0_model">
          {% for m in cn_models %}
            <option value="{{m}}" {% if p.cn0_model==m %}selected{% endif %}>{{m}}</option>
          {% endfor %}
        </select>
      </div>
      <div class="col">
        <label>Unit 0 Resize mode</label>
        <select name="cn0_resize">
          <option value="0" {% if p.cn0_resize==0 %}selected{% endif %}>Just Resize</option>
          <option value="1" {% if p.cn0_resize==1 %}selected{% endif %}>Crop and Resize</option>
          <option value="2" {% if p.cn0_resize==2 %}selected{% endif %}>Resize and Fill</option>
        </select>
      </div>
    </div>

    <!-- YENİ: Unit 0 Weight + Mode -->
    <div class="row">
      <div class="col">
        <label>Unit 0 Control Weight</label>
        <input type="number" name="cn0_weight" min="0" max="2" step="0.05" value="{{ p.cn0_weight or 0.5 }}">
      </div>
      <div class="col">
        <label>Unit 0 Control Mode</label>
        <select name="cn0_mode">
          <option value="0" {% if (p.cn0_mode or 0)==0 %}selected{% endif %}>Balanced</option>
          <option value="1" {% if (p.cn0_mode or 0)==1 %}selected{% endif %}>My prompt is more important</option>
          <option value="2" {% if (p.cn0_mode or 0)==2 %}selected{% endif %}>ControlNet is more important</option>
        </select>
      </div>
    </div>

    <div class="row">
      <div class="col">
        <label>Unit 1 Module</label>
        <select name="cn1_module">
          {% for m in cn_modules %}
            <option value="{{m}}" {% if p.cn1_module==m %}selected{% endif %}>{{m}}</option>
          {% endfor %}
        </select>
      </div>
      <div class="col">
        <label>Unit 1 Model</label>
        <select name="cn1_model">
          {% for m in cn_models %}
            <option value="{{m}}" {% if p.cn1_model==m %}selected{% endif %}>{{m}}</option>
          {% endfor %}
        </select>
      </div>
      <div class="col">
        <label>Unit 1 Resize mode</label>
        <select name="cn1_resize">
          <option value="0" {% if p.cn1_resize==0 %}selected{% endif %}>Just Resize</option>
          <option value="1" {% if p.cn1_resize==1 %}selected{% endif %}>Crop and Resize</option>
          <option value="2" {% if p.cn1_resize==2 %}selected{% endif %}>Resize and Fill</option>
        </select>
      </div>
    </div>

    <!-- YENİ: Unit 1 Weight + Mode -->
    <div class="row">
      <div class="col">
        <label>Unit 1 Control Weight</label>
        <input type="number" name="cn1_weight" min="0" max="2" step="0.05" value="{{ p.cn1_weight or 0.5 }}">
      </div>
      <div class="col">
        <label>Unit 1 Control Mode</label>
        <select name="cn1_mode">
          <option value="0" {% if (p.cn1_mode or 0)==0 %}selected{% endif %}>Balanced</option>
          <option value="1" {% if (p.cn1_mode or 0)==1 %}selected{% endif %}>My prompt is more important</option>
          <option value="2" {% if (p.cn1_mode or 0)==2 %}selected{% endif %}>ControlNet is more important</option>
        </select>
      </div>
    </div>


    <h4>Poz (Sayfaya özel)</h4>
    <div class="row">
      <div class="col" style="flex:1 1 100%">
        <label>Poz (dosya veya klasör yolu)</label>
        <input type="text" name="pose_path" value="{{ p.pose_path or '' }}">
        <div class="muted">Boşsa kitap ayarındaki varsayılan <code>poses_dir</code> kullanılır.</div>
      </div>
    </div>

    <h4>REActor (FaceSwap)</h4>
    <div class="row">
      <div class="col"><label>REActor Kullan</label><div class="switch"><input type="checkbox" name="use_reactor" {% if p.use_reactor %}checked{% endif %}></div></div>
    </div>
    <div class="row">
      <div class="col" style="flex:1 1 100%">
        <label>REActor Ayarları (JSON, opsiyonel)</label>
        <textarea name="reactor_json">{{ p.reactor_json or '' }}</textarea>
        <div class="muted">Örnek: <code>{"model":"inswapper_128.onnx","face_index":0}</code></div>
      </div>
    </div>

    <div class="btnrow"><button class="btn ok" type="submit">💾 Kaydet</button><a class="btn" href="{{ url_for('ui_book_pages', book_id=b.id) }}">İptal</a></div>
  </div></form>
{% endblock %}
"""

def default_page(next_index: int) -> dict:
    return {
        "id": uuid.uuid4().hex[:12],
        "index": next_index,
        "prompt": "",
        "negative_prompt": "",
        "checkpoint": "",
        "sampling_method": "Euler a",
        "sampling_steps": 12,
        "width": 1980,
        "height": 1020,
        "cfg_scale": 4.0,
        "seed": -1,
        "use_controlnet": True,

        # ControlNet (Instant-ID)
        "cn0_module": "InsightFace (InstantID)",
        "cn0_model": "ip-adapter_instant_id_sdxl [eb2d3ec0]",
        "cn0_resize": 1,
        "cn1_module": "instant_id_face_keypoints",
        "cn1_model": "control_instant_id_sdxl [c5c25a50]",
        "cn1_resize": 2,

        # YENİ: Control Weight + Control Mode (0=Balanced, 1=My prompt…, 2=ControlNet…)
        "cn0_weight": 0.5,
        "cn1_weight": 0.5,
        "cn0_mode": 0,
        "cn1_mode": 0,

        "styles": [],

        # --- Sayfa bazlı POSE ---
        "pose_path": "",

        # --- REActor ---
        "use_reactor": False,
        "reactor_json": "",

        "created_at": now_iso(),
        "updated_at": now_iso(),
    }


@app.route("/")
def root(): return redirect(url_for("ui_list_books"))

@app.route("/books")
def ui_list_books():
    return render_template_string(BOOK_LIST_HTML, books=list_books(), title=APP_TITLE)

def empty_book():
    return {
        "id": "",
        "name": "",
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "settings": {
            "data_source": "excel",
            "output_root": DEFAULT_OUT_DIR,
            "excel_path": "",
            "col_photo": "@photo",
            "col_first": "student name",
            "col_last": "student last name",
            "col_class": "class",
            "faces_dir": DEFAULT_FACES_DIR,
            "poses_dir": DEFAULT_POSES_DIR,   # fallback
            "col_out": "out",
        },
        "pages": []
    }

@app.route("/books/new", methods=["GET", "POST"])
def ui_new_book():
    b = empty_book()
    if request.method == "POST":
        b["id"] = uuid.uuid4().hex[:12]
        b["name"] = request.form.get("name", "").strip()
        s = b["settings"]
        s["data_source"] = request.form.get("data_source", s["data_source"])
        s["output_root"] = (request.form.get("output_root", s["output_root"]) or DEFAULT_OUT_DIR).strip()
        s["excel_path"] = request.form.get("excel_path", "").strip()
        s["col_photo"] = request.form.get("col_photo", s["col_photo"]).strip() or "@photo"
        s["col_first"] = request.form.get("col_first", s["col_first"]).strip()
        s["col_last"] = request.form.get("col_last", s["col_last"]).strip()
        s["col_class"] = request.form.get("col_class", s["col_class"]).strip()
        s["faces_dir"] = (request.form.get("faces_dir", s["faces_dir"]) or DEFAULT_FACES_DIR).strip()
        s["poses_dir"] = (request.form.get("poses_dir", s["poses_dir"]) or DEFAULT_POSES_DIR).strip()
        if not b["name"]:
            flash("Lütfen kitap adı girin.")
            return render_template_string(BOOK_FORM_HTML, b=b, title=APP_TITLE)
        write_book(b); flash("Kitap oluşturuldu.")
        return redirect(url_for("ui_book_pages", book_id=b["id"]))
    return render_template_string(BOOK_FORM_HTML, b=b, title=APP_TITLE)

@app.route("/books/<book_id>/edit", methods=["GET", "POST"])
def ui_edit_book(book_id):
    b = read_book(book_id)
    if not b: abort(404)
    if request.method == "POST":
        b["name"] = request.form.get("name", "").strip()
        s = b["settings"]
        s["data_source"] = request.form.get("data_source", s["data_source"])
        s["output_root"] = (request.form.get("output_root", s["output_root"]) or DEFAULT_OUT_DIR).strip()
        s["excel_path"] = request.form.get("excel_path", s["excel_path"]).strip()
        s["col_photo"] = request.form.get("col_photo", s["col_photo"]).strip() or "@photo"
        s["col_first"] = request.form.get("col_first", s["col_first"]).strip()
        s["col_last"] = request.form.get("col_last", s["col_last"]).strip()
        s["col_class"] = request.form.get("col_class", s["col_class"]).strip()
        s["faces_dir"] = (request.form.get("faces_dir", s["faces_dir"]) or DEFAULT_FACES_DIR).strip()
        s["poses_dir"] = (request.form.get("poses_dir", s["poses_dir"]) or DEFAULT_POSES_DIR).strip()
        b["updated_at"] = now_iso(); write_book(b); flash("Kitap ayarları güncellendi.")
        return redirect(url_for("ui_edit_book", book_id=book_id))
    return render_template_string(BOOK_FORM_HTML, b=b, title=APP_TITLE)

@app.route("/books/<book_id>/delete", methods=["POST"])
def ui_delete_book(book_id):
    p = book_file(book_id)
    if os.path.exists(p): os.remove(p); flash("Kitap silindi.")
    return redirect(url_for("ui_list_books"))

@app.route("/books/<book_id>/pages")
def ui_book_pages(book_id):
    b = read_book(book_id)
    if not b: abort(404)
    last = JOB_INDEX.get(book_id)
    status = JOBS.get(last, {}).get("status") if last else None
    return render_template_string(
        PAGES_HTML, b=b, title=f"{APP_TITLE} · {b['name']}",
        last_job_id=last, last_job_status=status or "-",
        models=api_models(), samplers=api_samplers(),
        cn_models=api_cn_model_list(), cn_modules=api_cn_module_list(),
        styles_list=api_styles()
    )

@app.route("/books/<book_id>/pages/add", methods=["POST"])
def ui_add_page(book_id):
    b = read_book(book_id)
    if not b: abort(404)
    p = default_page((len(b.get("pages", [])) + 1))
    p["index"] = int(request.form.get("index", p["index"]) or p["index"])
    p["seed"] = int(request.form.get("seed", p["seed"]) or p["seed"])
    p["use_controlnet"] = bool(request.form.get("use_controlnet"))
    p["checkpoint"] = request.form.get("checkpoint", p["checkpoint"]).strip()
    p["sampling_method"] = request.form.get("sampling_method", p["sampling_method"]).strip()
    p["sampling_steps"] = int(request.form.get("sampling_steps", p["sampling_steps"]) or p["sampling_steps"])
    p["width"] = max(64, min(2048, int(float(request.form.get("width", p["width"]) or p["width"]))))

    p["height"] = max(64, min(2048, int(float(request.form.get("height", p["height"]) or p["height"]))))

    p["cfg_scale"] = float(request.form.get("cfg_scale", p["cfg_scale"]) or p["cfg_scale"])
    p["prompt"] = request.form.get("prompt", "").strip()
    p["negative_prompt"] = request.form.get("negative_prompt", "").strip()
    p["cn0_module"] = request.form.get("cn0_module", p["cn0_module"]).strip()
    p["cn0_model"] = request.form.get("cn0_model", p["cn0_model"]).strip()
    p["cn0_resize"] = int(request.form.get("cn0_resize", p["cn0_resize"]))
    p["cn1_module"] = request.form.get("cn1_module", p["cn1_module"]).strip()
    p["cn1_model"] = request.form.get("cn1_model", p["cn1_model"]).strip()
    p["cn1_resize"] = int(request.form.get("cn1_resize", p["cn1_resize"]))
    p["styles"] = request.form.getlist("styles")
    # yeni: sayfa bazlı poz
    p["pose_path"] = request.form.get("pose_path", "").strip()
    # REActor
    p["use_reactor"] = bool(request.form.get("use_reactor"))
    p["reactor_json"] = request.form.get("reactor_json", "").strip()

    p["cn0_weight"] = float(request.form.get("cn0_weight", p.get("cn0_weight", 0.5)) or 0.5)
    p["cn1_weight"] = float(request.form.get("cn1_weight", p.get("cn1_weight", 0.5)) or 0.5)
    p["cn0_mode"]   = int(request.form.get("cn0_mode",   p.get("cn0_mode", 0)))
    p["cn1_mode"]   = int(request.form.get("cn1_mode",   p.get("cn1_mode", 0)))


    if not p["prompt"]:
        flash("Pozitif prompt boş olamaz.")
        return redirect(url_for("ui_book_pages", book_id=book_id))
    b.setdefault("pages", []).append(p)
    b["updated_at"] = now_iso(); write_book(b); flash("Sayfa eklendi.")
    return redirect(url_for("ui_book_pages", book_id=book_id))

@app.route("/books/<book_id>/pages/<page_id>/edit", methods=["GET", "POST"])
def ui_edit_page(book_id, page_id):
    b = read_book(book_id);
    if not b: abort(404)
    page = next((x for x in b.get("pages", []) if x["id"] == page_id), None)
    if not page: abort(404)
    if request.method == "POST":
        page["index"] = int(request.form.get("index", page["index"]) or page["index"])
        page["seed"] = int(request.form.get("seed", page["seed"]) or page["seed"])
        page["use_controlnet"] = bool(request.form.get("use_controlnet"))
        page["checkpoint"] = request.form.get("checkpoint", page["checkpoint"]).strip()
        page["sampling_method"] = request.form.get("sampling_method", page["sampling_method"]).strip()
        page["sampling_steps"] = int(request.form.get("sampling_steps", page["sampling_steps"]) or page["sampling_steps"])
        page["width"] = max(64, min(2048, int(float(request.form.get("width", page["width"]) or page["width"]))))

        page["height"] = max(64, min(2048, int(float(request.form.get("height", page["height"]) or page["height"]))))

        page["cfg_scale"] = float(request.form.get("cfg_scale", page["cfg_scale"]) or page["cfg_scale"])
        page["prompt"] = request.form.get("prompt", page["prompt"]).strip()
        page["negative_prompt"] = request.form.get("negative_prompt", page["negative_prompt"]).strip()
        page["cn0_module"] = request.form.get("cn0_module", page["cn0_module"]).strip()
        page["cn0_model"] = request.form.get("cn0_model", page["cn0_model"]).strip()
        page["cn0_resize"] = int(request.form.get("cn0_resize", page.get("cn0_resize", 1)))
        page["cn1_module"] = request.form.get("cn1_module", page["cn1_module"]).strip()
        page["cn1_model"] = request.form.get("cn1_model", page["cn1_model"]).strip()
        page["cn1_resize"] = int(request.form.get("cn1_resize", page.get("cn1_resize", 2)))
        page["styles"] = request.form.getlist("styles")

        page["cn0_weight"] = float(request.form.get("cn0_weight", page.get("cn0_weight", 0.5)) or 0.5)
        page["cn1_weight"] = float(request.form.get("cn1_weight", page.get("cn1_weight", 0.5)) or 0.5)
        page["cn0_mode"]   = int(request.form.get("cn0_mode",   page.get("cn0_mode", 0)))
        page["cn1_mode"]   = int(request.form.get("cn1_mode",   page.get("cn1_mode", 0)))

        # yeni: sayfa bazlı poz
        page["pose_path"] = request.form.get("pose_path", page.get("pose_path","")).strip()
        # REActor
        page["use_reactor"] = bool(request.form.get("use_reactor"))
        page["reactor_json"] = request.form.get("reactor_json", page.get("reactor_json","")).strip()
        page["updated_at"] = now_iso()
        b["updated_at"] = now_iso(); write_book(b); flash("Sayfa güncellendi.")
        return redirect(url_for("ui_book_pages", book_id=book_id))
    return render_template_string(PAGE_EDIT_HTML,
                                  b=b, p=page, title=f"{APP_TITLE} · Sayfa #{page['index']}",
                                  models=api_models(), samplers=api_samplers(),
                                  cn_models=api_cn_model_list(), cn_modules=api_cn_module_list(),
                                  styles_list=api_styles()
                                  )

@app.route("/books/<book_id>/pages/<page_id>/delete", methods=["POST"])
def ui_delete_page(book_id, page_id):
    b = read_book(book_id)
    if not b: abort(404)
    b["pages"] = [p for p in b.get("pages", []) if p["id"] != page_id]
    b["updated_at"] = now_iso(); write_book(b); flash("Sayfa silindi.")
    return redirect(url_for("ui_list_books"))

# ---------------- Prompt değişkenleri ----------------
def _make_key_variants(key: str) -> List[str]:
    """
    Bir sütun adı için çoklu anahtar varyantları üretir:
    - Orijinal / lower
    - Boşluklar '_' ve tamamen silinmiş (snake / nospace)
    - Türkçe katlama: ı->i, İ->I (hem orijinal hem lower formları)
    - Aksan/diakritik temizlenmiş ASCII formlar (ör. 'Adı' -> 'Adi')
    Bu varyantların hepsi için snake/nospace türevleri de eklenir.
    """
    k = (key or "").strip()
    if not k:
        return []

    def _forms(s: str) -> List[str]:
        snake   = re.sub(r"\s+", "_", s)
        nospace = re.sub(r"\s+", "", s)
        return [s, s.lower(), snake, snake.lower(), nospace, nospace.lower()]

    variants = set(_forms(k))

    # Türkçe 'ı/İ' katlaması
    tr = k.replace("ı", "i").replace("İ", "I")
    variants.update(_forms(tr))
    variants.update(_forms(tr.lower()))

    # Aksan/diakritik temizleme (örn. 'Adı' -> 'Adi')
    try:
        import unicodedata
        def _deaccent(s: str) -> str:
            return "".join(c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c))
        de = _deaccent(k)
        variants.update(_forms(de))
        # Türkçe katlanmış + deaccent
        de_tr = _deaccent(tr)
        variants.update(_forms(de_tr))
    except Exception:
        pass

    return [v for v in variants if v]


def _lower_tr(s: str) -> str:
    """Türkçe küçük harf dönüşümü (ı -> i, İ -> i)."""
    return (s or "").lower().replace("ı", "i").replace("İ", "i")

def _gender_en(raw: str) -> str:
    """Türkçe veya İngilizce cinsiyet değerini 'girl'/'boy' olarak normalize eder."""
    v = _lower_tr(raw).strip()
    if v in {"kiz", "kız", "k", "female", "f", "kadin", "kadın", "girl"}:
        return "girl"
    if v in {"erkek", "e", "male", "m", "boy", "adam"}:
        return "boy"
    return raw or ""


# Boşluk, Türkçe latin, @ ve _,- içerir (sütun adı boşluklu olabilir)
VAR_TOKEN_RE = re.compile(r"\{([A-Za-z0-9_ @\-\u00C0-\u024F\u1E00-\u1EFF]+)\}")

def render_text_template(text: str, child: Dict[str,Any]) -> str:
    """
    {KolonAdi} yer tutucularını, child['vars'] içindeki değerlerle doldurur.
    - Kolon adları case-insensitive.
    - 'Cinsiyet' özel kuralı: kız -> girl, erkek -> boy (tüm varyant anahtarlara uygulanır).
    - Sütun adını farklı şekillerde yazabil (örn. {student name}, {Student_Name}, {studentname}).
    """
    if not text:
        return text

    raw_vars: Dict[str, str] = {}

    # Baz yerleşikler
    raw_vars.update({
        "name": child.get("name",""),
        "class": child.get("class",""),
        "photo": child.get("face",""),
        "@photo": (child.get("vars", {}) or {}).get("@photo",""),
    })

    # Tüm satır değişkenlerini, birden fazla anahtar varyantıyla ekle
    row_vars = (child.get("vars") or {})
    for k, v in row_vars.items():
        val = "" if v is None else str(v)
        for key_variant in _make_key_variants(k):
            raw_vars[key_variant] = val

    # Cinsiyet özel haritalama (girl/boy). Olası başlık isimlerinin hepsini yakala.
    gender_keys = set()
    for cand in ["Cinsiyet", "cinsiyet", "Gender", "gender", "GENDER"]:
        gender_keys.update(_make_key_variants(cand))
    # ilk bulunan değeri al
    gender_raw = None
    for gk in gender_keys:
        if gk in raw_vars and raw_vars[gk]:
            gender_raw = raw_vars[gk]
            break
    if gender_raw is not None:
        mapped = _gender_en(gender_raw)  # kız->girl, erkek->boy, diğer -> raw
        for gk in gender_keys:
            raw_vars[gk] = mapped

    # Şablon doldurma
    def repl(m):
        key = (m.group(1) or "").strip()
        # Anahtar için varyantları üret ve sırayla dene
        for kv in _make_key_variants(key):
            if kv in raw_vars:
                return raw_vars[kv]
        # Birebir yazılan ad çakışırsa
        if key in raw_vars:
            return raw_vars[key]
        if key.lower() in raw_vars:
            return raw_vars[key.lower()]
        # bulunamazsa olduğu gibi bırak
        return m.group(0)

    return VAR_TOKEN_RE.sub(repl, text)



# ---------------- SD API yardımcıları ----------------
def read_image_to_b64(path: str) -> str:
    with Image.open(path) as im:
        buf = io.BytesIO()
        im.save(buf, format="PNG")
        return base64.b64encode(buf.getvalue()).decode("utf-8")

def b64_to_image(b64_str: str) -> Image.Image:
    return Image.open(io.BytesIO(base64.b64decode(b64_str))).copy()

def call_txt2img(payload: Dict[str, Any]) -> List[Image.Image]:
    r = requests.post(SD_BASE + "/sdapi/v1/txt2img", json=payload, timeout=120)
    r.raise_for_status()
    data = r.json()
    out = []
    for b64 in data.get("images", []):
        img_bytes = base64.b64decode(b64.split(",", 1)[-1])
        out.append(Image.open(io.BytesIO(img_bytes)).copy())
    return out

def build_controlnet_args(face_b64: str, pose_b64: Optional[str],
                          use_cnet: bool,
                          cn0_module: str, cn0_model: str, cn0_resize: int,
                          cn1_module: str, cn1_model: str, cn1_resize: int,
                          cn0_weight: float = 0.5, cn1_weight: float = 0.5,
                          cn0_control_mode: int = 0, cn1_control_mode: int = 0) -> Dict[str, Any]:
    """
    control_mode: 0=Balanced, 1=My prompt is more important, 2=ControlNet is more important
    """
    if not use_cnet:
        return {"args": []}

    unit0 = {
        "enabled": True, "module": cn0_module, "model": cn0_model,
        "weight": float(cn0_weight), "control_mode": int(cn0_control_mode),
        "image": face_b64, "resize_mode": int(cn0_resize),
        "guidance_start": 0.0, "guidance_end": 1.0,
        "pixel_perfect": False
    }
    u1_img = pose_b64 if pose_b64 else face_b64
    unit1 = {
        "enabled": True, "module": cn1_module, "model": cn1_model,
        "weight": float(cn1_weight), "control_mode": int(cn1_control_mode),
        "image": u1_img, "resize_mode": int(cn1_resize),
        "guidance_start": 0.0, "guidance_end": 1.0,
        "pixel_perfect": False
    }
    return {"args": [unit0, unit1]}


# ---------- REActor yardımcıları (POST-PROCESS) ----------
def pil_to_b64(img: Image.Image) -> str:
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return base64.b64encode(buf.getvalue()).decode("utf-8")

def _to_data_url(b64_plain: str) -> str:
    return b64_plain if b64_plain.startswith("data:image") else "data:image/png;base64," + b64_plain

def reactor_available() -> bool:
    try:
        r = requests.get(SD_BASE + "/reactor/models", timeout=5)
        return r.ok
    except Exception:
        return False

def reactor_swap(face_b64_plain: str, target_img: Image.Image, opts: Optional[dict] = None) -> Image.Image:
    t_b64_plain = pil_to_b64(target_img)

    # Varsayılanları UI’a yakın yap
    opts = opts or {}
    model = opts.get("model", "inswapper_128.onnx")
    face_index = int(opts.get("face_index", -1))             # <- otomatik
    source_face_index = int(opts.get("source_face_index", -1))# <- otomatik
    upscaler = opts.get("upscaler", "None")
    scale = int(opts.get("scale", 1))
    face_restorer = opts.get("face_restorer", "None")         # "CodeFormer" önerilir
    restorer_visibility = float(opts.get("restorer_visibility", 0.8 if face_restorer!="None" else 0))

    payload = {
        "source_image": _to_data_url(face_b64_plain),
        "target_image": _to_data_url(t_b64_plain),
        "source_faces_index": [source_face_index],
        "face_index": [face_index],
        "upscaler": upscaler,
        "scale": scale,
        "upscale_visibility": 1,
        "face_restorer": face_restorer,
        "restorer_visibility": restorer_visibility,
        "restore_first": 0,
        "model": model,
        "save_to_file": 0,
        "result_file_path": "",
    }

    r = requests.post(SD_BASE + "/reactor/image", json=payload, timeout=180)
    r.raise_for_status()
    js = r.json()
    out = js.get("image")
    if not out:
        raise RuntimeError("REActor boş döndü")
    if out.startswith("data:image"):
        out = out.split(",", 1)[-1]
    return b64_to_image(out)


# ---------- Excel/CSV out yazıcı ----------
class ExcelOutWriter:
    """
    - 'out' sütununa ; ile birleştirilmiş yolları yazar (eski davranış).
    - Ek olarak @sayfaN kolonlarını başlıkta GİZLİ TIRNAK ile üretir:  "'@sayfa1"
    - Yeni: @sayfaN başlıklarını mevcut başlıktaki **ilk boş sütundan** itibaren açar.
    """
    def __init__(self, path: str, col_out: Optional[str] = None, pages_prefix: str = "@sayfa"):
        self.path = path
        self.col_out = (col_out or "").strip()   # boş ise kapalı
        self.pages_prefix = (pages_prefix or "@sayfa").strip()
        self.ext = os.path.splitext(path)[1].lower()
        self.mode = "xlsx" if self.ext in (".xlsx", ".xlsm") else "csv"

        if self.mode == "xlsx":
            from openpyxl import load_workbook
            self.wb = load_workbook(path)
            self.ws = self.wb.active

            header_cells = list(self.ws.iter_rows(min_row=1, max_row=1))[0]
            self.header = [("" if c.value is None else str(c.value)) for c in header_cells]

            # out kolonu
            self.out_col_idx = None
            for i, h in enumerate(self.header, start=1):
                if str(h).strip().lower() == self.col_out.lower():
                    self.out_col_idx = i; break
            if self.out_col_idx is None:
                # out'u da ilk boş sütuna koy
                first_empty = self._first_empty_header_col_xlsx()
                self.out_col_idx = first_empty
                self.ws.cell(row=1, column=self.out_col_idx, value=self.col_out)

            # @sayfaN kolonları
            self.page_col_indices: Dict[int, int] = {}
            self._scan_existing_page_cols_xlsx()

        else:
            # CSV
            with open(path, newline="", encoding="utf-8-sig") as f:
                rows = list(csv.reader(f))
            if not rows: rows = [[]]
            self.header = rows[0]

            # out kolonu
            found = None
            for i, h in enumerate(self.header):
                if str(h).strip().lower() == self.col_out.lower():
                    found = i; break
            if found is None:
                idx = self._first_empty_header_col_csv()
                if idx is None:
                    self.header.append(self.col_out); found = len(self.header) - 1
                else:
                    # yerinde ekle
                    self.header[idx:idx+1] = [self.col_out]
                    found = idx
            self.out_col_idx = found
            self.rows = rows[1:]

            self.page_col_indices: Dict[int, int] = {}
            self._scan_existing_page_cols_csv()

    # ----- yardımcılar -----
    def _first_empty_header_col_xlsx(self) -> int:
        """1-based: başlıktaki ilk boş hücre sütun indexi; yoksa son + 1."""
        maxc = self.ws.max_column
        for c in range(1, maxc + 1):
            val = self.ws.cell(row=1, column=c).value
            if val is None or str(val).strip() == "":
                return c
        return maxc + 1

    def _first_empty_header_col_csv(self) -> Optional[int]:
        for i, h in enumerate(self.header):
            if (h is None) or (str(h).strip() == ""):
                return i
        return None

    def _scan_existing_page_cols_xlsx(self):
        for idx in range(1, self.ws.max_column + 1):
            s = self.ws.cell(row=1, column=idx).value
            if not isinstance(s, str): continue
            label = s.lstrip()
            if label.startswith("'@"): label = "@" + label.split("@",1)[1]
            if label.startswith(self.pages_prefix):
                try:
                    n = int(label[len(self.pages_prefix):]); self.page_col_indices[n] = idx
                except: pass

    def _scan_existing_page_cols_csv(self):
        for i, s in enumerate(self.header):
            s = (s or "")
            label = s.lstrip()
            if label.startswith("'@"): label = "@" + label.split("@",1)[1]
            if label.startswith(self.pages_prefix):
                try:
                    n = int(label[len(self.pages_prefix):]); self.page_col_indices[n] = i
                except: pass

    def _ensure_page_cols(self, max_n: int):
        if max_n <= 0: return
        if self.mode == "xlsx":
            # Başlangıç sütunu: mevcut başlıktaki ilk boş hücre
            start_col = self._first_empty_header_col_xlsx()
            # Mevcut sayfa kolonları varsa onların en küçüğünü başlangıç kabul edelim
            if self.page_col_indices:
                start_col = min(self.page_col_indices.values())
            col = start_col
            for n in range(1, max_n + 1):
                if n in self.page_col_indices:
                    continue
                # gizli tırnaklı başlık
                self.ws.cell(row=1, column=col, value=f"'@{self.pages_prefix.split('@',1)[-1]}{n}")
                self.page_col_indices[n] = col
                col += 1
        else:
            # CSV tarafında başlıkta ilk boş index
            start_idx = self._first_empty_header_col_csv()
            if self.page_col_indices:
                start_idx = min(self.page_col_indices.values())
            if start_idx is None:
                start_idx = len(self.header)
            # gerekli kadar başlık yerleştir
            needed = []
            for n in range(1, max_n + 1):
                if n not in self.page_col_indices:
                    needed.append((n, f"'@{self.pages_prefix.split('@',1)[-1]}{n}"))
            # araya yerleştir
            self.header[start_idx:start_idx] = [h for _, h in needed]
            # index haritasını kur
            for k, h in needed:
                pos = start_idx
                # her eklemeden sonra pos ilerler
                self.page_col_indices[k] = pos
                start_idx += 1
            # satır uzunluklarını yeni başlığa uydur
            for i in range(len(self.rows)):
                if len(self.rows[i]) < len(self.header):
                    self.rows[i] += [""] * (len(self.header) - len(self.rows[i]))

    # ----- public API -----
    def set_for_row(self, row_index_2based: int, value: str):
        if not row_index_2based or row_index_2based < 2: return
        if self.mode == "xlsx":
            self.ws.cell(row=row_index_2based, column=self.out_col_idx, value=value)
        else:
            i = row_index_2based - 2
            while i >= len(self.rows):
                self.rows.append([])
            if len(self.rows[i]) < len(self.header):
                self.rows[i] += [""] * (len(self.header) - len(self.rows[i]))
            self.rows[i][self.out_col_idx] = value

    def set_pages_for_row(self, row_index_2based: int, page_paths: List[str]):
        if not row_index_2based or row_index_2based < 2: return
        self._ensure_page_cols(len(page_paths))
        if self.mode == "xlsx":
            for n, p in enumerate(page_paths, start=1):
                col = self.page_col_indices.get(n)
                if col is not None:
                    self.ws.cell(row=row_index_2based, column=col, value=p)
        else:
            i = row_index_2based - 2
            while i >= len(self.rows):
                self.rows.append([])
            if len(self.rows[i]) < len(self.header):
                self.rows[i] += [""] * (len(self.header) - len(self.rows[i]))
            for n, p in enumerate(page_paths, start=1):
                col = self.page_col_indices.get(n)
                if col is not None:
                    self.rows[i][col] = p

    def save(self):
        if self.mode == "xlsx":
            self.wb.save(self.path)
        else:
            with open(self.path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(self.header)
                for r in self.rows:
                    if len(r) < len(self.header):
                        r = r + [""] * (len(self.header) - len(r))
                    w.writerow(r)

# ---- Kaynak okuyucular ----
def collect_children(settings: dict, log) -> List[Dict[str, str]]:
    """
    Kaynak çocuk listesini (yüz yolu, ad, sınıf, satır sırası, satır değişkenleri) döndürür.
    - settings:
        data_source: "excel" | "folders"
        faces_dir  : kök klasör (göreli @photo ile birleşir)
        excel_path : xlsx/xlsm/csv
        col_photo, col_first, col_last, col_class
    """
    children: List[Dict[str, str]] = []
    source = (settings.get("data_source") or "excel").strip().lower()
    faces_root = (settings.get("faces_dir") or "").strip()

    # ---- Yardımcılar ----
    def _normpath(p: str) -> str:
        return os.path.normpath(p) if p else p

    def _join_face(rel_or_abs: str) -> str:
        """Göreli @photo -> faces_root ile birleşir; abs ise olduğu gibi normalize edilir."""
        if not rel_or_abs:
            return ""
        rel_or_abs = str(rel_or_abs).replace("/", os.sep).replace("\\", os.sep).strip()
        if os.path.isabs(rel_or_abs):
            return _normpath(rel_or_abs)
        if not faces_root:
            # faces_root yoksa göreli yolu olduğu gibi kullan; yine normalize et
            return _normpath(rel_or_abs)
        return _normpath(os.path.join(faces_root, rel_or_abs))

    def _set_defaults_rowvars(row_vars: Dict[str, str], name: str, cls: str, rel_photo: str):
        row_vars.setdefault("name", name)
        row_vars.setdefault("class", cls)
        row_vars.setdefault("@photo", rel_photo)

    # ---- Klasör modu ----
    if source == "folders":
        exts = (".png", ".jpg", ".jpeg", ".webp", ".bmp")
        if not faces_root or not os.path.isdir(faces_root):
            log(f"[WARN] faces_dir klasörü bulunamadı: {faces_root}")
            return []
        row_i = 2  # Excel düzeni ile tutarlılık için 2’den başlatıyoruz
        for root, _, files in os.walk(faces_root):
            for n in files:
                if n.lower().endswith(exts):
                    face_abs = _normpath(os.path.join(root, n))
                    name = os.path.splitext(os.path.basename(n))[0]
                    cls = os.path.basename(os.path.dirname(face_abs))
                    rel_photo = os.path.relpath(face_abs, faces_root)
                    children.append({
                        "name": name,
                        "class": cls,
                        "face": face_abs,
                        "vars": {"name": name, "class": cls, "@photo": rel_photo},
                        "row_index": row_i
                    })
                    row_i += 1
        # Klasör modunda zorunlu bir sıra yok; alfabetik isimle hafif deterministik hale getirelim
        children.sort(key=lambda x: (x.get("class",""), x.get("name","")))
        return children

    # ---- Excel/CSV modu ----
    excel_path = (settings.get("excel_path") or "").strip()
    col_photo = (settings.get("col_photo") or "@photo").strip()
    col_first = (settings.get("col_first") or "").strip()
    col_last  = (settings.get("col_last")  or "").strip()
    col_class = (settings.get("col_class") or "").strip()

    if not excel_path or not os.path.exists(excel_path):
        log(f"[WARN] Excel yolu bulunamadı: {excel_path}")
        return []

    ext = os.path.splitext(excel_path)[1].lower()

    try:
        # ---- CSV ----
        if ext == ".csv":
            # Güvenli kodlama açılışı: utf-8-sig -> utf-8 -> cp1254 -> latin-1
            enc_trials = ("utf-8-sig", "utf-8", "cp1254", "latin-1")
            last_err = None
            for enc in enc_trials:
                try:
                    with open(excel_path, newline="", encoding=enc, errors="strict") as f:
                        reader = csv.DictReader(f)
                        # Başlıkları case-insensitive erişilebilsin diye her satırda hem orijinal hem lower tutacağız
                        for i, row in enumerate(reader, start=2):
                            # DictReader anahtarları dosyadaki başlıklarla gelir
                            row_raw = (row or {})
                            # Tüm varyant anahtarlarla satır sözlüğü
                            row_dict: Dict[str, str] = {}
                            for k, v in row_raw.items():
                                kk = (k or "").strip()
                                vv = "" if v is None else str(v).strip()
                                for key_variant in _make_key_variants(kk):
                                    row_dict[key_variant] = vv

                            # İstenen kolonları case-insensitive al
                            def get_ci(key: str, default: str = "") -> str:
                                if not key: return default
                                return row_dict.get(key, row_dict.get(key.lower(), default))

                            photo_rel = get_ci(col_photo) or get_ci("@photo") or get_ci("photo")
                            if not photo_rel:
                                # Foto yoksa satırı atla
                                continue

                            rel = str(photo_rel).replace("/", os.sep).replace("\\", os.sep)
                            face_abs = _join_face(rel)

                            if not os.path.exists(face_abs):
                                log(f"[WARN] Yüz dosyası yok: {face_abs}")
                                continue

                            first = get_ci(col_first, "")
                            last  = get_ci(col_last, "")
                            cls   = get_ci(col_class, "")
                            name  = " ".join([x for x in [first, last] if x]).strip() or os.path.splitext(os.path.basename(face_abs))[0]

                            _set_defaults_rowvars(row_dict, name, cls, rel)
                            children.append({
                                "name": name,
                                "class": cls,
                                "face": face_abs,
                                "vars": row_dict,
                                "row_index": i
                            })
                    # Başarılı açılış → döngü kır
                    last_err = None
                    break
                except Exception as e:
                    last_err = e
                    continue
            if last_err:
                raise last_err

        # ---- XLSX / XLSM / (XLS de openpyxl ile kısıtlı) ----
        else:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active

            # Başlık satırı
            header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
            header = [("" if c is None else str(c)) for c in header]

            # Case-insensitive index haritası
            idx_ci = {}
            for i, h in enumerate(header):
                h_str = (h or "").strip()
                if h_str:
                    idx_ci[h_str] = i
                    idx_ci[h_str.lower()] = i

            def get_cell(row_tuple, col_name: str, default: str = "") -> str:
                if not col_name:
                    return default
                j = idx_ci.get(col_name, idx_ci.get(col_name.lower()))
                if j is None or j >= len(row_tuple):
                    return default
                v = row_tuple[j]
                return "" if v is None else str(v).strip()

            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # Satırı dict'e çevir (tüm varyant anahtarlarla)
                row_dict: Dict[str, str] = {}
                for h, j in ((h, idx_ci.get(h)) for h in header if h in idx_ci):
                    if j is None or j >= len(row):
                        continue
                    val = row[j]
                    vv = "" if val is None else str(val).strip()
                    hh = (h or "").strip()
                    for key_variant in _make_key_variants(hh):
                        row_dict[key_variant] = vv

                photo_rel = get_cell(row, col_photo) or get_cell(row, "@photo") or get_cell(row, "photo")
                if not photo_rel:
                    continue

                rel = str(photo_rel).replace("/", os.sep).replace("\\", os.sep)
                face_abs = _join_face(rel)

                if not os.path.exists(face_abs):
                    log(f"[WARN] Yüz dosyası yok: {face_abs}")
                    continue

                first = get_cell(row, col_first, "")
                last  = get_cell(row, col_last, "")
                cls   = get_cell(row, col_class, "")
                name  = " ".join([x for x in [first, last] if x]).strip() or os.path.splitext(os.path.basename(face_abs))[0]

                _set_defaults_rowvars(row_dict, name, cls, rel)
                children.append({
                    "name": name,
                    "class": cls,
                    "face": face_abs,
                    "vars": row_dict,
                    "row_index": i
                })

    except Exception as e:
        log(f"[ERR] Excel/CSV okuma hatası: {e}")

    # ---- Dönüş öncesi: yol normalizasyonu + sıraya sok ----
    for ch in children:
        ch["face"] = _normpath(ch.get("face", ""))
    # Excel sırası garanti: row_index küçükten büyüğe
    children.sort(key=lambda x: x.get("row_index", 10**9))
    return children


# ---- POSE bulucu ----
def find_pose_image_path(poses_dir_or_file: str) -> Optional[str]:
    if not poses_dir_or_file: return None
    p = poses_dir_or_file.strip()
    if os.path.isfile(p):
        return p
    if os.path.isdir(p):
        exts = (".png", ".jpg", ".jpeg", ".webp", ".bmp")
        files = [os.path.join(p, n) for n in os.listdir(p) if n.lower().endswith(exts)]
        files.sort()
        if files: return files[0]
    return None

# ---- Çalıştırma ----
def run_book_via_api(book: dict, log_path: str, out_dir: str = DEFAULT_OUT_DIR, progress_cb=None):
    name = book["name"]
    s = book["settings"]
    out_root = s.get("output_root") or out_dir
    pages = sorted(book.get("pages", []), key=lambda p: p.get("index", 0))
    os.makedirs(out_root, exist_ok=True)

    def log(msg: str):
        print(msg)
        with open(log_path, "a", encoding="utf-8") as f: f.write(msg.rstrip() + "\n")

    children = collect_children(s, log)

    # POSE cache
    pose_cache: Dict[str, Optional[str]] = {}
    def resolve_pose_b64(pose_source: str) -> (Optional[str], Optional[str]):
        if not pose_source: return (None, None)
        if pose_source in pose_cache:
            resolved = find_pose_image_path(pose_source)
            return (resolved, pose_cache[pose_source])
        resolved = find_pose_image_path(pose_source)
        if resolved and os.path.exists(resolved):
            try:
                b64 = read_image_to_b64(resolved)
                pose_cache[pose_source] = b64
                return (resolved, b64)
            except Exception as e:
                log(f"[WARN] Poz okunamadı: {resolved} ({e})")
        pose_cache[pose_source] = None
        return (resolved, None)

    # EXCEL out yazıcı
    writer = None
    if s.get("data_source") == "excel" and s.get("excel_path") and os.path.exists(s["excel_path"]):
        try:
            writer = ExcelOutWriter(s["excel_path"], col_out=None)  # out tamamen kapalı

        except Exception as e:
            log(f"[WARN] Excel out yazıcı açılamadı: {e}")

    log(f"[INFO] Başlıyor: {name} | children:{len(children)} pages:{len(pages)}")
    if not children:
        log("[WARN] Kaynakta çocuk bulunamadı.")
        log("[DONE] Tamamlandı."); return

    # REActor hazır mı?
    reactor_ok = reactor_available()
    if not reactor_ok:
        log("[REACTOR] endpoint yok (Forge/A1111'da REActor eklentisi etkin mi?).")

    # Yardımcı: sayfa çıktı dosya yolu
    def page_output_path(base_dir: str, page_index: int) -> str:
        # Alt klasör yok; doğrudan sayfa{N}.png
        os.makedirs(base_dir, exist_ok=True)
        return os.path.join(base_dir, f"sayfa{int(page_index)}.png")

    for child in children:
        child_name  = (child.get("name")  or "").strip()
        child_class = (child.get("class") or "").strip()
        face_path   = child["face"]

        # Bu çocuğun baz çıkış klasörü
        base_out = os.path.join(out_root, child_class) if child_class else out_root
        child_out = os.path.join(base_out, child_name)
        os.makedirs(child_out, exist_ok=True)

        # Hangi sayfalar bitmiş?
        done_map = {}
        for p in pages:
            p_idx = int(p.get("index", 0) or 0)
            out_p = page_output_path(child_out, p_idx)
            done_map[p_idx] = os.path.exists(out_p)

        all_done = all(done_map.values()) if done_map else False
        if all_done:
            log(f"[SKIP] {child_name} | tüm sayfalar mevcut, atlanıyor.")
            # Excel 'out' sütununu mevcut dosyalarla da güncelleyelim (varsa)
            if writer and child.get("row_index"):
                try:
                    existing = [page_output_path(child_out, int(p.get('index',0) or 0)) for p in pages]
                    existing = [p for p in existing if os.path.exists(p)]
                    if existing:
                        #writer.set_for_row(child["row_index"], "; ".join(existing))
                        writer.set_pages_for_row(child["row_index"], existing)  # ← @sayfaN kolonları
                        writer.save()
                        log(f"[EXCEL] out (skip) güncellendi (satır {child['row_index']}): {existing[-1]}")
                except Exception as e:
                    log(f"[WARN] Excel out (skip) yazılamadı: {e}")
            continue

        try:
            face_b64 = read_image_to_b64(face_path)
        except Exception as e:
            log(f"[WARN] Yüz okunamadı: {face_path} ({e})"); continue

        log(f"[CHILD] {child_name} | class={child_class or '-'} | face={face_path}")

        out_paths_for_child: List[str] = []
        # Mevcut olanları listeye ekle (Excel için)
        for p in pages:
            p_idx = int(p.get("index", 0) or 0)
            out_p = page_output_path(child_out, p_idx)
            if os.path.exists(out_p):
                out_paths_for_child.append(out_p)

        for p in pages:
            p_idx = int(p.get("index", 0) or 0)
            out_p = page_output_path(child_out, p_idx)

            if os.path.exists(out_p):
                log(f"[SKIP] Page {p_idx} zaten var → {out_p}")
                continue

            seed = int(p.get("seed", -1))
            pr = render_text_template(p.get("prompt", ""), child)
            npr = render_text_template(p.get("negative_prompt", ""), child)

            # Sayfa bazlı poz → boşsa kitap ayarı fallback
            pose_source = (p.get("pose_path") or s.get("poses_dir") or "").strip()
            pose_used_path, pose_b64 = resolve_pose_b64(pose_source) if pose_source else (None, None)
            if pose_used_path and pose_b64:
                log(f"[POSE] Page {p_idx} → {pose_used_path}")
            else:
                log(f"[POSE] Page {p_idx} → (yok)")

            log(f"[PAGE] {p_idx} | seed={seed} | {p.get('width')}x{p.get('height')} | steps={p.get('sampling_steps')}")

            # ControlNet
            cn = build_controlnet_args(
                face_b64=face_b64,
                pose_b64=pose_b64,
                use_cnet=bool(p.get("use_controlnet", True)),
                cn0_module=p.get("cn0_module", "InsightFace (InstantID)"),
                cn0_model=p.get("cn0_model",  "ip-adapter_instant_id_sdxl [eb2d3ec0]"),
                cn0_resize=int(p.get("cn0_resize",1)),
                cn1_module=p.get("cn1_module", "instant_id_face_keypoints"),
                cn1_model=p.get("cn1_model",  "control_instant_id_sdxl [c5c25a50]"),
                cn1_resize=int(p.get("cn1_resize",2)),
                cn0_weight=float(p.get("cn0_weight", 0.5)),
                cn1_weight=float(p.get("cn1_weight", 0.5)),
                cn0_control_mode=int(p.get("cn0_mode", 0)),
                cn1_control_mode=int(p.get("cn1_mode", 0))
            )


            log(f"[CN] u0_module='{p.get('cn0_module')}' u0_model='{p.get('cn0_model')}' resize={int(p.get('cn0_resize',1))}; "
                f"u1_module='{p.get('cn1_module')}' u1_model='{p.get('cn1_model')}' resize={int(p.get('cn1_resize',2))}; "
                f"u1_image={'POSE' if pose_b64 else 'FACE'}")

            payload = {
                "prompt": pr, "negative_prompt": npr,
                "width": int(p.get("width", 1024)), "height": int(p.get("height", 1024)),
                "sampler_name": p.get("sampling_method", "Euler a"),
                "steps": int(p.get("sampling_steps", 20)),
                "cfg_scale": float(p.get("cfg_scale", 7.0)),
                "seed": seed,
                "override_settings": {"sd_model_checkpoint": p.get("checkpoint", "")},
                "alwayson_scripts": {"ControlNet": cn},
                "styles": p.get("styles", []),
            }

            try:
                imgs = call_txt2img(payload)
                if not imgs:
                    log("[WARN] API bir görüntü döndürmedi."); continue

                gen_img = imgs[0]

                # DEBUG CN input
                try:
                    b64_to_image(face_b64).save(os.path.join(os.path.dirname(out_p), "debug_cn0_input.png"))
                    (b64_to_image(pose_b64) if pose_b64 else b64_to_image(face_b64)).save(os.path.join(os.path.dirname(out_p), "debug_cn1_input.png"))
                except Exception:
                    pass

                # --- REActor (dış API ile post-process) ---
                if p.get("use_reactor") and reactor_ok:
                    reactor_opts = {}
                    rj_text = p.get("reactor_json", "").strip()
                    if rj_text:
                        try:
                            rj = json.loads(rj_text)
                            if isinstance(rj, dict):
                                for key in ("model","face_index","source_face_index","upscaler","scale",
                                            "upscale_visibility","face_restorer","restorer_visibility",
                                            "restore_first","gender_source","gender_target"):
                                    if key in rj:
                                        reactor_opts[key] = rj[key]
                        except Exception as e:
                            log(f"[REACTOR] JSON yok sayıldı (parse): {e}")
                    try:
                        gen_img = reactor_swap(face_b64, gen_img, reactor_opts)
                        log("[REACTOR] swap uygulandı.")
                    except Exception as e:
                        log(f"[REACTOR] başarısız, orijinal kullanılacak: {e}")

                gen_img.save(out_p)
                out_paths_for_child.append(out_p)
                log(f"[OK] Kaydedildi: {out_p}")
                if callable(progress_cb):
                    progress_cb({"event":"save","image_path":out_p,"child":child_name,"class":child_class,"page_index":p_idx})
            except Exception as e:
                log(f"[ERR] API hata: {e}")

        # Çocuk tamamlandı → Excel 'out' yaz
        if writer and child.get("row_index"):
            try:
                #writer.set_for_row(child["row_index"], "; ".join(out_paths_for_child))
                writer.set_pages_for_row(child["row_index"], out_paths_for_child)  # ← @sayfaN kolonları
                writer.save()
                if out_paths_for_child:
                    log(f"[EXCEL] out yazıldı (satır {child['row_index']}): {out_paths_for_child[-1]}")
            except Exception as e:
                log(f"[WARN] Excel out yazılamadı: {e}")

    if writer:
        try: writer.save()
        except: pass

    log("[DONE] Tamamlandı.")




# ---- İş yönetimi + SSE ----
JOBS: Dict[str, Dict[str, Any]] = {}
JOB_INDEX: Dict[str, str] = {}

def start_job(book_id):
    job_id = uuid.uuid4().hex[:12]
    log_path = os.path.join(LOGS_DIR, f"{job_id}.log")
    JOBS[job_id] = {"status": "running", "log_path": log_path, "book_id": book_id, "started_at": now_iso(),
                    "finished_at": None, "last_image": None, "last_child": None, "last_page": None, "kind": "api"}
    JOB_INDEX[book_id] = job_id
    def worker():
        try:
            def progress_cb(info):
                JOBS[job_id]["last_image"] = info.get("image_path")
                JOBS[job_id]["last_child"] = info.get("child")
                JOBS[job_id]["last_page"] = info.get("page_index")
            run_book_via_api(read_book(book_id), log_path=log_path, out_dir=DEFAULT_OUT_DIR, progress_cb=progress_cb)
            JOBS[job_id]["status"] = "finished"
        except Exception as e:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n[ERR] {e}\n")
            JOBS[job_id]["status"] = "failed"
        finally:
            JOBS[job_id]["finished_at"] = now_iso()
    threading.Thread(target=worker, daemon=True).start()
    return job_id

# === Forge UI Üzerinden Çalıştır ===
IMG_PATH_RE = re.compile(r"""(?ix)
    (?:Kaydedildi:\s*|[() ]*çıktı:\s*)
    (
      (?:[A-Z]:\\|/)?      # Windows sürücüsü ya da kök / opsiyonel
      [^\n\r]+?\.
      (?:png|jpg|jpeg|webp|bmp)
    )
""")

def start_job_ui(book_id, forge_url: str):
    """runner_ui_prompts.py --book-id <id> --forge-url <forge_url> --batch [--children-json manifest]"""
    if not os.path.exists(RUNNER_PATH):
        raise RuntimeError(f"UI runner bulunamadı: {RUNNER_PATH}")

    job_id = uuid.uuid4().hex[:12]
    log_path = os.path.join(LOGS_DIR, f"{job_id}.log")
    JOBS[job_id] = {
        "status": "running",
        "log_path": log_path,
        "book_id": book_id,
        "started_at": now_iso(),
        "finished_at": None,
        "last_image": None,
        "last_child": None,
        "last_page": None,
        "kind": "ui"
    }
    JOB_INDEX[book_id] = job_id

    # --- Excel sırası: manifest oluştur ---
    book = read_book(book_id)
    if not book:
        raise RuntimeError(f"Kitap bulunamadı: {book_id}")
    settings = book.get("settings", {}) or {}

    # API ile aynı toplama/sıralama mantığı
    ordered_children = collect_children(settings, log=lambda *_: None)  # row_index'e göre sıralı döner
    manifest_path = os.path.join(LOGS_DIR, f"{job_id}_children.json")
    with open(manifest_path, "w", encoding="utf-8") as mf:
        json.dump(ordered_children, mf, ensure_ascii=False, indent=2)

    args = [
        sys.executable, RUNNER_PATH,
        "--book-id", book_id,
        "--forge-url", forge_url,
        "--batch",
        "--children-json", manifest_path  # ← kritik: Excel sırası runner'a aktarılıyor
    ]
    if RUNNER_HEADLESS:
        args.append("--headless")
    if "--keep-open" not in args:
        args.append("--keep-open")

    # Alt süreç IO'sunu UTF-8'de sabitle
    env = os.environ.copy()
    env["PYTHONIOENCODING"] = "utf-8"
    env["PYTHONUTF8"] = "1"

    def worker():
        try:
            with open(log_path, "w", encoding="utf-8") as lf:
                lf.write(f"[INFO] UI runner başlatılıyor: {' '.join(args)}\n")
                lf.write(f"[INFO] Children manifest: {manifest_path}\n")
                lf.write(f"[INFO] Çocuk sayısı (Excel sırası): {len(ordered_children)}\n")
                if ordered_children:
                    preview_names = ", ".join([c.get('name','?') for c in ordered_children[:5]])
                    lf.write(f"[INFO] İlk 5: {preview_names}\n")

            proc = subprocess.Popen(
                args,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                bufsize=1,
                env=env
            )

            with open(log_path, "a", encoding="utf-8") as lf:
                for line in proc.stdout:
                    s = line.rstrip("\n")
                    lf.write(s + "\n")
                    # Mevcut önizleme yakalama mantığınız burada kalabilir.
                    m = IMG_PATH_RE.search(s)
                    if m:
                        path = m.group(1).strip()
                        if os.path.exists(path):
                            JOBS[job_id]["last_image"] = path
                            try:
                                pi = re.search(r"Sayfa\s*#(\d+)", s, re.I)
                                if pi:
                                    JOBS[job_id]["last_page"] = int(pi.group(1))
                            except:
                                pass

            proc.wait()

            # --- UI Çalıştıktan sonra: Excel'e @sayfaN + out yaz ---
            try:
                book2 = read_book(book_id)
                if book2:
                    settings2 = book2.get("settings", {}) or {}
                    out_root2 = settings2.get("output_root") or DEFAULT_OUT_DIR
                    pages2 = sorted(book2.get("pages", []), key=lambda p: int(p.get("index", 0) or 0))

                    # Excel Writer sadece excel kaynağı varsa açılır
                    writer2 = None
                    if settings2.get("data_source") == "excel" and settings2.get("excel_path") and os.path.exists(settings2["excel_path"]):
                        writer2 = ExcelOutWriter(settings2["excel_path"], col_out=settings2.get("col_out", "out"))

                    def page_output_path2(base_dir: str, page_index: int) -> str:
                        os.makedirs(base_dir, exist_ok=True)
                        return os.path.join(base_dir, f"sayfa{int(page_index)}.png")

                    if writer2:
                        # Çocukları Excel sırasıyla gez
                        ordered_children2 = collect_children(settings2, log=lambda *_: None)
                        for ch in ordered_children2:
                            child_name  = (ch.get("name")  or "").strip()
                            child_class = (ch.get("class") or "").strip()
                            base_out = os.path.join(out_root2, child_class) if child_class else out_root2
                            child_out = os.path.join(base_out, child_name)

                            # Bu çocuk için üretilmiş tüm sayfaları topla
                            paths = []
                            for p in pages2:
                                p_idx = int(p.get("index", 0) or 0)
                                out_p = page_output_path2(child_out, p_idx)
                                if os.path.exists(out_p):
                                    paths.append(out_p)

                            if paths and ch.get("row_index"):
                                # out: ; ile birleştirilmiş
                                #writer2.set_for_row(ch["row_index"], "; ".join(paths))
                                # '@sayfaN' kolonları
                                writer2.set_pages_for_row(ch["row_index"], paths)

                        writer2.save()
                        with open(log_path, "a", encoding="utf-8") as lf:
                            lf.write("[EXCEL] UI işlemi sonrası @sayfaN + out yazıldı.\n")
            except Exception as e:
                with open(log_path, "a", encoding="utf-8") as lf:
                    lf.write(f"[WARN] UI sonrası Excel yazılamadı: {e}\n")

            JOBS[job_id]["status"] = "finished" if proc.returncode == 0 else "failed"
        except Exception as e:
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"\n[ERR] {e}\n")
            JOBS[job_id]["status"] = "failed"
        finally:
            JOBS[job_id]["finished_at"] = now_iso()

    threading.Thread(target=worker, daemon=True).start()
    return job_id


@app.route("/books/<book_id>/run", methods=["POST"])
def ui_run_book(book_id):
    if not read_book(book_id): abort(404)
    job_id = start_job(book_id); flash(f"İş (API) başlatıldı: {job_id[:8]}")
    return redirect(url_for("ui_book_pages", book_id=book_id))

@app.route("/books/<book_id>/run-ui", methods=["POST"])
def ui_run_book_ui(book_id):
    b = read_book(book_id)
    if not b: abort(404)
    forge_url = SD_BASE
    try:
        job_id = start_job_ui(book_id, forge_url)
        flash(f"İş (Forge UI) başlatıldı: {job_id[:8]}")
    except Exception as e:
        flash(f"Hata: {e}")
    return redirect(url_for("ui_book_pages", book_id=book_id))

@app.route("/jobs/<job_id>")
def ui_job_status(job_id):
    j = JOBS.get(job_id)
    if not j: abort(404)
    log = ""
    if os.path.exists(j["log_path"]):
        with open(j["log_path"], "r", encoding="utf-8") as f:
            log = f.read()
    html = f"""
    {{% extends "base.html" %}}{{% block content %}}
      <div class="panel">
        <h2>İş: {job_id}</h2>
        <p>Durum: <span id="job-status" class="status">{j['status']}</span></p>
        <p>Tür: <code>{j.get('kind','api')}</code></p>
        <p>Kitap: <a href="{{{{ url_for('ui_book_pages', book_id='{j['book_id']}') }}}}">{j['book_id']}</a></p>
        <div class="row">
          <div class="col" style="min-width:320px;flex:2 1 520px">
            <label>Log</label>
            <pre id="logbox" style="white-space: pre-wrap; background:#0b0e12; padding:12px; border-radius:10px; border:1px solid #1f2732; max-height:60vh; overflow:auto;">{log}</pre>
          </div>
          <div class="col" style="min-width:260px;flex:1 1 260px">
            <label>Son Önizleme</label>
            <div class="preview-wrap"
                 style="position:relative;border:1px solid #1f2732;border-radius:10px;background:#0b0e12;
                        padding:8px;display:flex;align-items:center;justify-content:center;min-height:240px;">
              <img id="preview" src="" alt="preview" style="max-width:100%; max-height:420px; display:none;" />
              <div id="ppct"
                   style="position:absolute;right:10px;top:10px;font-size:12px;background:rgba(0,0,0,.55);
                          padding:2px 6px;border-radius:6px;display:none;">0%</div>
              <div id="pbar"
                   style="position:absolute;left:0;bottom:0;height:6px;width:0%;
                          background:#3b82f6;border-bottom-left-radius:10px;border-bottom-right-radius:10px;
                          transition:width .2s;"></div>
            </div>
          </div>
        </div>
      </div>
      <script>
        (function(){{
          const logbox   = document.getElementById('logbox');
          const statusEl = document.getElementById('job-status');
          const img      = document.getElementById('preview');
          const pbar     = document.getElementById('pbar');
          const ppct     = document.getElementById('ppct');
          const es = new EventSource("{{{{ url_for('job_stream', job_id='{job_id}') }}}}");
          const previewUrl = "{{{{ url_for('job_preview', job_id='{job_id}') }}}}";
          es.addEventListener('log', e => {{
            logbox.textContent += (logbox.textContent.endsWith("\\n") ? "" : "\\n") + e.data;
            logbox.scrollTop = logbox.scrollHeight;
          }});
          es.addEventListener('image', e => {{
            img.src = previewUrl + "?ts=" + e.data;
            img.style.display = 'block';
          }});
          es.addEventListener('progress', e => {{
            const v = Math.max(0, Math.min(100, parseInt(e.data || '0', 10)));
            pbar.style.width = v + '%';
            ppct.textContent = v + '%';
            if (v > 0) ppct.style.display = 'block';
            if (v >= 100) setTimeout(() => {{ ppct.style.display = 'none'; }}, 800);
          }});
          es.addEventListener('done', e => {{
            statusEl.textContent = e.data;
            pbar.style.width = '100%';
            setTimeout(() => {{ ppct.style.display = 'none'; }}, 800);
            es.close();
          }});
        }})();
      </script>
    {{% endblock %}}
    """
    return render_template_string(html, title=f"İş · {job_id}")

@app.route("/jobs/<job_id>/stream")
def job_stream(job_id):
    j = JOBS.get(job_id)
    if not j: abort(404)
    path = j["log_path"]
    kind = j.get("kind","api")
    def generate():
        while not os.path.exists(path):
            if JOBS.get(job_id, {}).get("status") != "running": break
            time.sleep(0.2)
        last_img_ts = 0
        last_prog = -1
        try:
            with open(path, "r", encoding="utf-8") as f:
                f.seek(0, os.SEEK_END)
                while True:
                    line = f.readline()
                    if line:
                        yield f"event: log\ndata: {line.rstrip()}\n\n"
                    else:
                        li = JOBS.get(job_id, {}).get("last_image")
                        if li and os.path.exists(li):
                            ts = int(os.path.getmtime(li))
                            if ts != last_img_ts:
                                last_img_ts = ts
                                yield f"event: image\ndata: {ts}\n\n"
                        if kind == "api":
                            try:
                                resp = requests.get(SD_BASE + "/sdapi/v1/progress?skip_current_image=true", timeout=2)
                                if resp.ok:
                                    pj = resp.json() or {}
                                    pct = int(round(float(pj.get("progress") or 0.0) * 100))
                                    if pct != last_prog:
                                        last_prog = pct
                                        yield f"event: progress\ndata: {pct}\n\n"
                            except Exception:
                                pass
                        st = JOBS.get(job_id, {}).get("status")
                        if st != "running":
                            if last_prog < 100:
                                yield "event: progress\ndata: 100\n\n"
                            yield f"event: done\ndata: {st}\n\n"
                            break
                        time.sleep(0.5)
        except GeneratorExit:
            return
    return Response(stream_with_context(generate()), mimetype="text/event-stream")

@app.route("/jobs/<job_id>/preview")
def job_preview(job_id):
    j = JOBS.get(job_id)
    if not j: abort(404)
    p = j.get("last_image")
    if not p or not os.path.exists(p): abort(404)
    resp = send_file(p, mimetype="image/png"); resp.headers["Cache-Control"] = "no-store"
    return resp

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=5055, debug=True)
