# runner_ui_prompts.py
# Otomasyon: faces_dir altındaki öğrenciler -> her öğrenci için tüm sayfalar
# Prompt/Neg + Steps/Width/Height/CFG/Seed + Sampler + Styles + Generate
# REActor upload (stale-safe) + ControlNet (Unit0/Unit1)
# ÇIKTI: İşlenmiş görsel (galeri/canvas/img) üretim bittikten sonra alınıp
# output_root/SINIF/AD SOYAD/sayfaX/image.png olarak kaydedilir.
# pip install selenium webdriver-manager pandas openpyxl

import os, json, argparse, time, glob, re, base64, hashlib
from pathlib import Path

import pandas as pd
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import StaleElementReferenceException
from webdriver_manager.chrome import ChromeDriverManager


ROOT_DIR   = Path(__file__).resolve().parent
DATA_DIR   = ROOT_DIR / "data"
BOOKS_DIR  = DATA_DIR / "books"

IMG_PATTERNS = ("*.png","*.jpg","*.jpeg","*.webp","*.bmp")




# ------------------------ IO ------------------------
def load_book(book_id: str) -> dict:
    p = BOOKS_DIR / f"{book_id}.json"
    if not p.exists():
        raise FileNotFoundError(f"Book JSON bulunamadı: {p}")
    with open(p, "r", encoding="utf-8") as f:
        return json.load(f)

def pick_page(book: dict, page_index: int | None, page_id: str | None) -> dict:
    pages = list(book.get("pages", []))
    if not pages:
        raise ValueError("Kitapta hiç sayfa yok.")
    if page_id:
        for p in pages:
            if p.get("id") == page_id:
                return p
        raise ValueError(f"page_id bulunamadı: {page_id}")
    if not page_index or page_index < 1:
        return sorted(pages, key=lambda x: x.get("index", 0))[0]
    for p in pages:
        if int(p.get("index", 0)) == int(page_index):
            return p
    pages.sort(key=lambda x: x.get("index", 0))
    return pages[min(len(pages)-1, page_index-1)]

def list_faces_in_dir(faces_dir: str | None) -> list[str]:
    if not faces_dir: return []
    hits = []
    for patt in IMG_PATTERNS:
        hits.extend(glob.glob(os.path.join(faces_dir, "**", patt), recursive=True))
    hits = [os.path.abspath(x) for x in hits]
    hits.sort(key=lambda p: (Path(p).parent.as_posix().lower(), Path(p).name.lower()))
    return hits

def ensure_dir(p: str | Path):
    Path(p).mkdir(parents=True, exist_ok=True)


# --------------------- Selenium base ----------------
def new_driver(headless: bool = False):
    from selenium.webdriver.chrome.options import Options
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--start-maximized")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1600,1000")
    opts.add_argument("--lang=en-US")
    drv = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=opts)
    return drv

def to_fullscreen(driver):
    # YALNIZCA pencereyi büyüt (F11 yok)
    try:
        driver.maximize_window()
    except:
        pass

def _get_app(driver):
    els = driver.find_elements(By.CSS_SELECTOR, "gradio-app")
    return els[0] if els else None

def query_one(driver, selector: str):
    app = _get_app(driver)
    if app:
        el = driver.execute_script("""
          const app = arguments[0], sel = arguments[1];
          const root = app.shadowRoot ? app.shadowRoot : app;
          return root.querySelector(sel);
        """, app, selector)
        if el: return el
    try:    return driver.find_element(By.CSS_SELECTOR, selector)
    except: return None

def query_all(driver, selector: str):
    app = _get_app(driver)
    if app:
        els = driver.execute_script("""
          const app = arguments[0], sel = arguments[1];
          const root = app.shadowRoot ? app.shadowRoot : app;
          return root.querySelectorAll(sel);
        """, app, selector)
        if els: return els
    try:    return driver.find_elements(By.CSS_SELECTOR, selector)
    except: return []

def fire(driver, el):
    driver.execute_script("""
      const el = arguments[0];
      ['input','change','blur'].forEach(t => el.dispatchEvent(new Event(t,{bubbles:true})));
    """, el)

def set_text(driver, el, val: str):
    driver.execute_script("arguments[0].value = arguments[1];", el, val or "")
    fire(driver, el)

def set_number(driver, el, val):
    try:
        n = int(val) if isinstance(val, (int, float, str)) and str(val).strip() != "" else ""
    except:
        n = ""
    driver.execute_script("arguments[0].value = arguments[1];", el, n); fire(driver, el)

def set_float(driver, el, val):
    try:
        x = float(val) if isinstance(val, (int, float, str)) and str(val).strip() != "" else ""
    except:
        x = ""
    driver.execute_script("arguments[0].value = arguments[1];", el, x); fire(driver, el)

def set_select_by_text_or_value(driver, sel_el, target):
    if not sel_el or target is None: return False
    target_s = str(target).strip().lower()
    opts = sel_el.find_elements(By.CSS_SELECTOR, "option")
    chosen = None
    for o in opts:
        tv = (o.get_attribute("value") or "").strip().lower()
        tt = (o.text or "").strip().lower()
        if tv == target_s or tt == target_s:
            chosen = o; break
    if not chosen:
        for o in opts:
            tt = (o.text or "").strip().lower()
            if target_s and target_s in tt:
                chosen = o; break
    if chosen:
        Select(sel_el).select_by_value(chosen.get_attribute("value"))
        fire(driver, sel_el)
        return True
    return False


# -------- prompts helpers --------
def find_prompt_textareas(driver):
    el_box  = query_one(driver, "#txt2img_prompt textarea") or query_one(driver, "#txt2img_prompt label textarea")
    neg_box = query_one(driver, "#txt2img_neg_prompt textarea") or query_one(driver, '#txt2img_neg_prompt label textarea')
    if el_box and neg_box:
        return el_box, neg_box
    prompts = query_all(driver, 'textarea[placeholder^="Prompt"]')
    negs    = query_all(driver, 'textarea[placeholder^="Negative prompt"]')
    def pick_visible(cands):
        for el in cands:
            try:
                if el.is_displayed(): return el
            except: pass
        return cands[0] if cands else None
    return pick_visible(prompts), pick_visible(negs)

def maybe_switch_to_txt2img(driver):
    for bt in query_all(driver, "button"):
        try:
            t = (bt.text or "").strip().lower()
            if "txt2img" in t or "text to image" in t:
                bt.click(); time.sleep(0.2); return True
        except: pass
    return False


# -------- listbox helper --------
def _get_listbox_value(driver, inp):
    val = (inp.get_attribute("value") or "").strip()
    if not val:
        val = (inp.get_attribute("title") or "").strip()
    return val

def open_dropdown_and_pick(driver, aria_label: str, target_text: str, timeout=3):
    if not target_text: return False
    target_text = str(target_text).strip()
    inp = query_one(driver, f'input[role="listbox"][aria-label="{aria_label}"]')
    if not inp:
        wrap = query_one(driver, f'div[aria-label="{aria_label}"] .wrap') or query_one(driver, 'div:has(> .wrap)')
        if (wrap):
            try:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", wrap)
                wrap.click(); time.sleep(0.1)
            except: pass
        inp = query_one(driver, f'input[role="listbox"][aria-label="{aria_label}"]')
        if not inp: return False

    cur = (_get_listbox_value(driver, inp) or "").lower()
    if cur == target_text.lower():
        return True

    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        inp.click(); time.sleep(0.05)
        inp.send_keys(Keys.CONTROL, "a")
        inp.send_keys(target_text)
        time.sleep(0.05)
        inp.send_keys(Keys.ENTER)
        time.sleep(0.12)
        cur2 = (_get_listbox_value(driver, inp) or "").lower()
        if cur2 == target_text.lower():
            return True
    except: pass

    try:
        inp.click(); time.sleep(0.1)
    except: pass
    parent = query_one(driver, "#dropdown-options")
    end = time.time() + timeout
    tgt = target_text.lower()
    while time.time() < end:
        opts = query_all(driver, '#dropdown-options [role="option"], #dropdown-options div')
        for o in opts:
            try:
                t = (o.text or "").strip().lower()
                if t and tgt in t:
                    o.click(); time.sleep(0.1)
                    cur3 = (_get_listbox_value(driver, inp) or "").lower()
                    return cur3 == tgt
            except: pass
        if parent:
            try: driver.execute_script("arguments[0].scrollTop = arguments[0].scrollTop + 240;", parent)
            except: pass
        time.sleep(0.05)

    try:
        driver.execute_script("arguments[0].value = arguments[1];", inp, target_text)
        fire(driver, inp)
        time.sleep(0.05)
        inp.send_keys(Keys.ENTER)
        time.sleep(0.1)
        cur4 = (_get_listbox_value(driver, inp) or "").lower()
        return cur4 == tgt
    except:
        return False


# ---------- Prompt yeniden yazma ----------
def get_txt2img_textareas(driver):
    p_el  = query_one(driver, "#txt2img_prompt textarea") or query_one(driver, '#txt2img_prompt label textarea')
    n_el  = query_one(driver, "#txt2img_neg_prompt textarea") or query_one(driver, '#txt2img_neg_prompt label textarea')
    if not p_el or not n_el:
        p_el = p_el or query_one(driver, 'textarea[placeholder^="Prompt"]')
        n_el = n_el or query_one(driver, 'textarea[placeholder^="Negative prompt"]')
    return p_el, n_el

def write_text_with_retry(driver, el, text, attempts=3, pause=0.2):
    last_err = None
    for _ in range(max(1, attempts)):
        try:
            driver.execute_script("arguments[0].value = arguments[1];", el, text or ""); fire(driver, el); time.sleep(pause)
            val = (el.get_attribute("value") or "")
            if (val or "").strip() == (text or "").strip(): return True
        except Exception as e:
            last_err = e; time.sleep(pause)
    if last_err: print("⚠️ prompt yazma tekrarında hata:", last_err)
    return False


# ----------------- REActor: tek tık açma ------------------
def _find_labelwrap_by_text(driver, text_contains: str):
    text_contains = (text_contains or "").strip().lower()
    for b in query_all(driver, "button.label-wrap"):
        try:
            t = (b.text or "").strip().lower()
            if text_contains in t:
                return b
        except: pass
    return None

def _is_section_open_via_icon(btn):
    try:
        icon = btn.find_element(By.CSS_SELECTOR, ".icon")
        style = (icon.get_attribute("style") or "").replace(" ", "").lower()
        if "rotate(0deg)" in style:  return True
        if "rotate(90deg)" in style: return False
    except: pass
    try:
        cls = btn.get_attribute("class") or ""
        return "open" in cls.split()
    except: return False

def wait_inputs_stable(driver, scope, css='input[type="file"]', min_count=1, stable_ms=350, timeout=8):
    end = time.time() + timeout
    last_cnt = -1
    last_change = time.time()
    while time.time() < end:
        try: els = scope.find_elements(By.CSS_SELECTOR, css)
        except Exception: els = []
        cnt = len(els); now = time.time()
        if cnt != last_cnt:
            last_cnt = cnt; last_change = now
        if cnt >= min_count and (now - last_change) * 1000.0 >= stable_ms:
            return True
        time.sleep(0.08)
    return False

def _find_reactor_button(driver):
    return _find_labelwrap_by_text(driver, "reactor")

def _reactor_container_from_button(driver, btn):
    if not btn: return None
    return driver.execute_script("""
        let el = arguments[0];
        let p = el.parentElement;
        for (let i=0;i<20 && p;i++){
            if (p.querySelector('input[type="file"]')) return p;
            p = p.parentElement;
        }
        return el.parentElement;
    """, btn)

def open_reactor_once(driver, wait_timeout=6):
    btn = _find_reactor_button(driver)
    if not btn: return False
    if _is_section_open_via_icon(btn):  # zaten açık
        return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn); time.sleep(0.05)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", btn),
        lambda: ActionChains(driver).move_to_element(btn).pause(0.03).click().perform(),
        lambda: btn.click(),
    ):
        try: fn(); break
        except: pass
    cont = _reactor_container_from_button(driver, btn)
    if cont and wait_inputs_stable(driver, cont, 'input[type="file"]', 1, 300, wait_timeout):
        return True
    return _is_section_open_via_icon(btn)

def _reactor_upload_confirmed(driver, container, file_input=None):
    # input.files kontrolü
    if file_input:
        try:
            files_len = driver.execute_script("return arguments[0] && arguments[0].files ? arguments[0].files.length : 0;", file_input)
            if (files_len or 0) > 0:
                return True
        except:
            pass
    if container:
        try:
            any_has_file = driver.execute_script("""
                const c = arguments[0];
                if (!c) return 0;
                const ins = c.querySelectorAll('input[type="file"]');
                for (const el of ins) {
                    if (el.files && el.files.length > 0) return 1;
                }
                return 0;
            """, container)
            if any_has_file:
                return True
        except:
            pass
    # görünür img/canvas
    try:
        img_cnt = driver.execute_script("""
            const c = arguments[0];
            if (!c) return 0;
            const nodes = c.querySelectorAll('img, canvas');
            let vis = 0;
            nodes.forEach(el=>{
                const st = getComputedStyle(el);
                if (st && st.display!=='none' && st.visibility!=='hidden' && (el.width||el.clientWidth)>0) vis++;
            });
            return vis;
        """, container) or 0
        return int(img_cnt) > 0
    except:
        return False

def upload_face_in_reactor_panel(driver, face_path: str):
    if not face_path or not os.path.exists(face_path):
        return False
    abs_path = os.path.abspath(face_path)
    if not open_reactor_once(driver): return False
    btn = _find_reactor_button(driver)
    if not btn: return False
    container = _reactor_container_from_button(driver, btn)
    if not container: return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", container)
        time.sleep(0.05)
    except: pass
    high_conf = None
    try:
        high_conf = driver.execute_script("""
            const cont = arguments[0];
            return cont.querySelector('button.svelte-j5bxrl input[type="file"][data-testid="file-upload"]')
                || cont.querySelector('input[type="file"][aria-label="file upload"]')
                || cont.querySelector('input[type="file"]');
        """, container)
    except: pass
    if not high_conf: return False
    if _reactor_upload_confirmed(driver, container, high_conf): return True
    try:
        driver.execute_script("""
            const el = arguments[0];
            el.style.display='block'; el.style.visibility='visible'; el.style.opacity=1;
            el.style.position='fixed'; el.style.left='10px'; el.style.top='10px';
            el.style.width='4px'; el.style.height='4px'; el.style.zIndex=2147483647;
        """, high_conf)
    except: pass
    try:
        try:
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", high_conf)
        except: pass
        time.sleep(0.02)
        high_conf.send_keys(abs_path)
        fire(driver, high_conf)
    except StaleElementReferenceException:
        if _reactor_upload_confirmed(driver, container, None): return True
    except Exception:
        if not _reactor_upload_confirmed(driver, container, None):
            return False
    end_local = time.time() + 9.0
    while time.time() < end_local:
        if _reactor_upload_confirmed(driver, container, high_conf):
            return True
        time.sleep(0.08)
    time.sleep(0.4)
    return _reactor_upload_confirmed(driver, container, high_conf)


# ----------------- ControlNet Integrated (Unit 0 / shared) -----------------
def scroll_to_bottom(driver, steps=8, pause=0.12):
    try:
        for _ in range(steps):
            driver.execute_script("window.scrollBy(0, document.documentElement.clientHeight * 0.75);")
            time.sleep(pause)
    except: pass

def _find_controlnet_button(driver):
    return _find_labelwrap_by_text(driver, "controlnet integrated")

def _is_controlnet_open(driver, btn):
    return _is_section_open_via_icon(btn)

def open_controlnet_once(driver):
    btn = _find_controlnet_button(driver)
    if not btn:
        scroll_to_bottom(driver, steps=10)
        btn = _find_controlnet_button(driver)
        if not btn:
            return False
    if _is_controlnet_open(driver, btn):
        return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn); time.sleep(0.05)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", btn),
        lambda: ActionChains(driver).move_to_element(btn).pause(0.03).click().perform(),
        lambda: btn.click(),
    ):
        try: fn(); break
        except: pass
    time.sleep(0.18)
    return _is_controlnet_open(driver, btn)

def ensure_controlnet_checkbox_on(driver):
    if not open_controlnet_once(driver):
        return False
    end = time.time() + 4.0
    cb = None
    while time.time() < end and not cb:
        cb = query_one(driver, 'input#input-accordion-1-visible-checkbox') or \
             query_one(driver, 'input.input-accordion-checkbox')
        if cb: break
        time.sleep(0.1)
    if not cb: return False
    try:
        checked = cb.is_selected()
    except:
        checked = (cb.get_attribute("checked") is not None)
    if checked: return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb); time.sleep(0.05)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", cb),
        lambda: ActionChains(driver).move_to_element(cb).pause(0.03).click().perform(),
        lambda: cb.click(),
    ):
        try: fn(); time.sleep(0.12); break
        except: pass
    try: return cb.is_selected()
    except: return (cb.get_attribute("checked") is not None)

def _controlnet_scope(driver):
    btn = _find_controlnet_button(driver)
    if not btn: return None
    scope = driver.execute_script("""
        let el = arguments[0];
        let p = el.parentElement;
        for (let i=0;i<25 && p;i++){
            if (p.querySelector('div.forge-image-container')) return p;
            p = p.parentElement;
        }
        return el.parentElement;
    """, btn)
    return scope or btn

def _controlnet_upload_confirmed(driver, container):
    return driver.execute_script("""
        const c = arguments[0];
        if (!c) return false;
        const img = c.querySelector('img.forge-image');
        if (img && img.src && getComputedStyle(img).display !== 'none') return true;
        const cnv = c.querySelector('canvas.forge-drawing-canvas');
        if (cnv && cnv.width > 0 && cnv.height > 0) return true;
        return false;
    """, container)

def _query_in_scope(driver, scope, css):
    try:
        return driver.execute_script("return arguments[0].querySelector(arguments[1]);", scope, css)
    except:
        return None

def _query_all_in_scope(driver, scope, css):
    try:
        return driver.execute_script("return arguments[0].querySelectorAll(arguments[1]);", scope, css)
    except:
        return []


def upload_student_image_to_controlnet(driver, img_path: str):
    if not img_path or not os.path.exists(img_path):
        return False
    if not open_controlnet_once(driver):
        return False
    scope = _controlnet_scope(driver)
    if not scope:
        return False
    container = _query_in_scope(driver, scope, 'div.forge-image-container')
    if not container:
        scroll_to_bottom(driver, steps=4)
        container = _query_in_scope(driver, scope, 'div.forge-image-container')
        if not container:
            return False
    file_inp = _query_in_scope(driver, container, 'input[type="file"]') or _query_in_scope(driver, scope, 'input[type="file"]')
    if not file_inp:
        return False
    abs_path = os.path.abspath(img_path)
    try:
        driver.execute_script("""
            const el = arguments[0];
            el.style.display='block'; el.style.visibility='visible'; el.style.opacity=1;
            el.style.position='fixed'; el.style.left='8px'; el.style.top='8px';
            el.style.width='3px'; el.style.height='3px'; el.style.zIndex=2147483647;
        """, file_inp)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", file_inp)
        file_inp.send_keys(abs_path); fire(driver, file_inp)
        end = time.time() + 6.0
        while time.time() < end:
            if _controlnet_upload_confirmed(driver, container):
                return True
            time.sleep(0.1)
    except Exception as e:
        print("controlnet upload fail:", e)
    return False


def tick_pixel_perfect(driver):
    if not open_controlnet_once(driver): return False
    scope = _controlnet_scope(driver)
    if not scope: return False
    labels = _query_all_in_scope(driver, scope, "label")
    target_cb = None
    for lb in labels:
        try:
            txt = (lb.text or "").strip().lower()
            if "pixel perfect" in txt:
                try:
                    cb = lb.find_element(By.CSS_SELECTOR, 'input[type="checkbox"]')
                except:
                    cb = _query_in_scope(driver, lb, 'input[type="checkbox"]')
                if cb: target_cb = cb; break
        except:
            continue
    if not target_cb:
        target_cb = _query_in_scope(driver, scope, 'input[type="checkbox"][data-testid="checkbox"]')
    if not target_cb: return False
    try:
        checked = target_cb.is_selected()
    except:
        checked = (target_cb.get_attribute("checked") is not None)
    if checked: return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", target_cb); time.sleep(0.05)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", target_cb),
        lambda: ActionChains(driver).move_to_element(target_cb).pause(0.03).click().perform(),
        lambda: target_cb.click(),
    ):
        try: fn(); time.sleep(0.08); break
        except: pass
    try: return target_cb.is_selected()
    except: return (target_cb.get_attribute("checked") is not None)

def select_instant_id_radio(driver):
    if not open_controlnet_once(driver): return False
    scope = _controlnet_scope(driver)
    if not scope: return False
    radio = _query_in_scope(driver, scope, 'input[type="radio"][value="Instant-ID"]')
    if not radio:
        radios = _query_all_in_scope(driver, scope, 'input[type="radio"]')
        for r in radios:
            text = ""
            try:
                lbl = r.find_element(By.XPATH, "./ancestor::label")
                text = (lbl.text or "").strip()
            except:
                try:
                    lbl = driver.execute_script("return arguments[0].closest('label');", r)
                    text = (lbl.text or "").strip() if lbl else ""
                except:
                    text = ""
            if text.lower().strip().replace(" ", "") in ("instant-id","instantid"):
                radio = r; break
    if not radio: return False
    try:
        checked = (radio.get_attribute("aria-checked") == "true") or radio.is_selected()
    except:
        checked = (radio.get_attribute("checked") is not None)
    if checked: return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", radio); time.sleep(0.04)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", radio),
        lambda: ActionChains(driver).move_to_element(radio).pause(0.02).click().perform(),
        lambda: radio.click(),
    ):
        try: fn(); time.sleep(0.06); break
        except: pass
    try:
        return (radio.get_attribute("aria-checked") == "true") or radio.is_selected()
    except:
        return (radio.get_attribute("checked") is not None)


# --- OpenPose radyo seçimi (ana ControlNet alanı) ---
def select_openpose_radio(driver):
    """
    Forge arayüzünde ControlNet Integrated panelindeki 'OpenPose' radio butonunu seçer.
    """
    try:
        radios = query_all(driver, 'input[type="radio"]')
        for r in radios:
            val = (r.get_attribute("value") or "").strip().lower()
            if val == "openpose":
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", r)
                time.sleep(0.05)
                if (r.get_attribute("aria-checked") != "true"):
                    r.click()
                    time.sleep(0.1)
                return True
        print("⚠️ OpenPose radio bulunamadı")
        return False
    except Exception as e:
        print("⚠️ OpenPose radio seçimi hata:", e)
        return False

# --- Resize mode yardımcıları (Integrated + Unit N) ---

from selenium.webdriver.common.by import By

_RESIZE_LABELS = {  # 0=Just, 1=Crop, 2=Fill
    0: "Just Resize",
    1: "Crop and Resize",
    2: "Resize and Fill",
}

def _click_radio_by_value_in(container, value_text: str) -> bool:
    if not container or not value_text:
        return False
    try:
        radios = container.find_elements(By.CSS_SELECTOR, 'input[type="radio"]')
    except Exception:
        # container driver ise:
        try:
            radios = container.find_elements(By.CSS_SELECTOR, 'input[type="radio"]')
        except Exception:
            radios = []
    tgt = (value_text or "").strip().lower()
    for r in radios:
        try:
            val = (r.get_attribute("value") or "").strip().lower()
            if val == tgt:
                try:
                    container.parent.execute_script("arguments[0].scrollIntoView({block:'center'});", r)
                except Exception:
                    pass
                time.sleep(0.04)
                if r.get_attribute("aria-checked") != "true":
                    r.click()
                    time.sleep(0.08)
                return True
        except Exception:
            continue
    return False

def _controlnet_integrated_container(driver):
    # ControlNet integrated alanının kapsayıcısı
    btn = _find_controlnet_button(driver)
    if not btn: return None
    try:
        return driver.execute_script("""
            let el = arguments[0]; let p = el.parentElement;
            for (let i=0;i<25 && p;i++){
                if (p.querySelector('div.forge-image-container')) return p;
                p = p.parentElement;
            }
            return el.parentElement;
        """, btn)
    except Exception:
        return None

def _set_resize_mode_integrated(driver, mode: int) -> bool:
    """Unit 0 (Integrated) için resize modu seçer."""
    if not open_controlnet_once(driver): return False
    cont = _controlnet_integrated_container(driver)
    value_text = _RESIZE_LABELS.get(int(mode), None)
    return _click_radio_by_value_in(cont or driver, value_text) if value_text else False

def _cn_unit_get_container(driver, unit_index: int):
    btn = _find_cn_unit_button(driver, unit_index)
    if not btn: return None
    try:
        return driver.execute_script("""
            let el = arguments[0]; let p = el.parentElement;
            for (let i=0;i<25 && p;i++){
                if (p.querySelector('div.forge-image-container')) return p;
                p = p.parentElement;
            }
            return el.parentElement;
        """, btn)
    except Exception:
        return None

def _cn_unit_select_resize_mode(driver, unit_index: int, mode: int) -> bool:
    """Unit N için 0=Just,1=Crop,2=Fill"""
    if not _open_cn_unit_once(driver, unit_index): return False
    cont = _cn_unit_get_container(driver, unit_index)
    value_text = _RESIZE_LABELS.get(int(mode), None)
    return _click_radio_by_value_in(cont or driver, value_text) if value_text else False



def _is_instant_module(name: str) -> bool:
    s = (name or "").strip().lower()
    # InsightFace (InstantID) ve instant_id_face_keypoints sinyalleri
    return ("insightface" in s) or ("instant_id_face_keypoints" in s)

def _is_openpose_module(name: str) -> bool:
    s = (name or "").strip().lower()
    # Kullanım: openpose_full seçildiyse true
    return "openpose_full" in s


# --- ControlNet: Preprocessor & Model (scope'lu dropdown) ---
def select_listbox_scoped(driver, scope, aria_label: str, target_text: str, timeout=3):
    if not target_text:
        return False
    target_text = str(target_text).strip()
    def _get_inp():
        inp = _query_in_scope(driver, scope, f'input[role="listbox"][aria-label="{aria_label}"]')
        if not inp:
            wrap = _query_in_scope(driver, scope, f'div[aria-label="{aria_label}"] .wrap')
            if wrap:
                try:
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", wrap)
                    wrap.click(); time.sleep(0.05)
                except:
                    pass
                inp = _query_in_scope(driver, scope, f'input[role="listbox"][aria-label="{aria_label}"]')
        return inp
    inp = _get_inp()
    if not inp:
        return False
    def get_val():
        v = (inp.get_attribute("value") or "").strip()
        if not v:
            v = (inp.get_attribute("title") or "").strip()
        return v
    if get_val().lower() == target_text.lower():
        return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
        inp.click(); time.sleep(0.03)
        inp.send_keys(Keys.CONTROL, "a"); inp.send_keys(target_text); time.sleep(0.03); inp.send_keys(Keys.ENTER)
        time.sleep(0.08)
        if get_val().lower() == target_text.lower():
            return True
    except:
        pass
    try:
        inp.click(); time.sleep(0.05)
    except:
        pass
    parent = query_one(driver, "#dropdown-options")
    end = time.time() + timeout
    tgt = target_text.lower()
    def try_pick(match_exact: bool):
        opts = query_all(driver, '#dropdown-options [role="option"], #dropdown-options div')
        for o in opts:
            try:
                t = (o.text or "").strip()
                tl = t.lower()
                ok = (tl == tgt) if match_exact else (tgt in tl)
                if ok:
                    o.click(); time.sleep(0.08)
                    return get_val().lower() == tgt
            except:
                continue
        return False
    while time.time() < end:
        if try_pick(True) or try_pick(False):
            return True
        if parent:
            try:
                driver.execute_script("arguments[0].scrollTop = arguments[0].scrollTop + 220;", parent)
            except:
                pass
        time.sleep(0.04)
    try:
        driver.execute_script("arguments[0].value = arguments[1];", inp, target_text)
        fire(driver, inp); time.sleep(0.04); inp.send_keys(Keys.ENTER); time.sleep(0.08)
        return get_val().lower() == tgt
    except:
        return False

def select_controlnet_preproc_and_model(driver, module_text: str | None, model_text: str | None):
    if not open_controlnet_once(driver): return False
    scope = _controlnet_scope(driver)
    if not scope: return False
    ok_any = False
    if module_text:
        if not select_listbox_scoped(driver, scope, "Preprocessor", module_text):
            print("⚠️ ControlNet Preprocessor seçilemedi:", module_text)
        else:
            ok_any = True
    if model_text:
        if not select_listbox_scoped(driver, scope, "Model", model_text):
            print("⚠️ ControlNet Model seçilemedi:", model_text)
        else:
            ok_any = True
    return ok_any


# ----------------- ControlNet Unit 1 (EKLENDİ) -----------------
def _find_cn_unit_button(driver, unit_index: int) -> object | None:
    return _find_labelwrap_by_text(driver, f"controlnet unit {unit_index}")

def _is_cn_unit_open(driver, unit_index: int) -> bool:
    btn = _find_cn_unit_button(driver, unit_index)
    if not btn: return False
    return _is_section_open_via_icon(btn)

def _open_cn_unit_once(driver, unit_index: int) -> bool:
    btn = _find_cn_unit_button(driver, unit_index)
    if not btn:
        scroll_to_bottom(driver, steps=12)
        btn = _find_cn_unit_button(driver, unit_index)
        if not btn: return False
    if _is_cn_unit_open(driver, unit_index): return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn); time.sleep(0.04)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", btn),
        lambda: ActionChains(driver).move_to_element(btn).pause(0.02).click().perform(),
        lambda: btn.click(),
    ):
        try: fn(); break
        except: pass
    time.sleep(0.12)
    return _is_cn_unit_open(driver, unit_index)

def _cn_unit_scope(driver, unit_index: int):
    btn = _find_cn_unit_button(driver, unit_index)
    if not btn: return None
    scope = driver.execute_script("""
        let el = arguments[0];
        let p = el.parentElement;
        for (let i=0;i<25 && p;i++){
            if (p.querySelector('div.forge-image-container')) return p;
            p = p.parentElement;
        }
        return el.parentElement;
    """, btn)
    return scope or btn

def ensure_cn_unit_checkbox_on(driver, unit_index: int) -> bool:
    if not _open_cn_unit_once(driver, unit_index): return False
    expect_id = f"input-accordion-{unit_index+1}-visible-checkbox"  # Unit 1 -> id: 2
    scope = _cn_unit_scope(driver, unit_index)
    cb = _query_in_scope(driver, scope, f'input#{expect_id}') or _query_in_scope(driver, scope, 'input.input-accordion-checkbox')
    if not cb: return False
    try:
        checked = cb.is_selected()
    except:
        checked = (cb.get_attribute("checked") is not None)
    if checked: return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", cb); time.sleep(0.04)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", cb),
        lambda: ActionChains(driver).move_to_element(cb).pause(0.02).click().perform(),
        lambda: cb.click(),
    ):
        try: fn(); time.sleep(0.08); break
        except: pass
    try: return cb.is_selected()
    except: return (cb.get_attribute("checked") is not None)

def cn_unit_upload_image(driver, unit_index: int, img_path: str) -> bool:
    if not img_path or not os.path.exists(img_path): return False
    if not _open_cn_unit_once(driver, unit_index): return False
    scope = _cn_unit_scope(driver, unit_index)
    if not scope: return False
    container = _query_in_scope(driver, scope, 'div.forge-image-container')
    if not container:
        scroll_to_bottom(driver, steps=3)
        container = _query_in_scope(driver, scope, 'div.forge-image-container')
        if not container: return False
    file_inp = _query_in_scope(driver, container, 'input[type="file"]') or _query_in_scope(driver, scope, 'input[type="file"]')
    if not file_inp: return False
    abs_path = os.path.abspath(img_path)
    try:
        driver.execute_script("""
            const el = arguments[0];
            el.style.display='block'; el.style.visibility='visible'; el.style.opacity=1;
            el.style.position='fixed'; el.style.left='8px'; el.style.top='8px';
            el.style.width='3px'; el.style.height='3px'; el.style.zIndex=2147483647;
        """, file_inp)
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", file_inp)
        file_inp.send_keys(abs_path); fire(driver, file_inp)
        end = time.time() + 6.0
        while time.time() < end:
            if _controlnet_upload_confirmed(driver, container):
                return True
            time.sleep(0.08)
    except Exception as e:
        print("cn unit upload fail:", e)
    return False

def cn_unit_select_instant_id(driver, unit_index: int) -> bool:
    if not _open_cn_unit_once(driver, unit_index): return False
    scope = _cn_unit_scope(driver, unit_index)
    if not scope: return False
    radio = _query_in_scope(driver, scope, 'input[type="radio"][value="Instant-ID"]')
    if not radio: return False
    try:
        checked = (radio.get_attribute("aria-checked") == "true") or radio.is_selected()
    except:
        checked = (radio.get_attribute("checked") is not None)
    if checked: return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", radio); time.sleep(0.03)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", radio),
        lambda: ActionChains(driver).move_to_element(radio).pause(0.02).click().perform(),
        lambda: radio.click(),
    ):
        try: fn(); time.sleep(0.06); break
        except: pass
    try:
        return (radio.get_attribute("aria-checked") == "true") or radio.is_selected()
    except:
        return (radio.get_attribute("checked") is not None)

def cn_unit_select_resize_and_fill(driver, unit_index: int) -> bool:
    if not _open_cn_unit_once(driver, unit_index): return False
    scope = _cn_unit_scope(driver, unit_index)
    if not scope: return False
    radio = None
    radios = _query_all_in_scope(driver, scope, 'input[type="radio"]')
    for r in radios:
        try:
            val = (r.get_attribute("value") or "").strip().lower()
            if val == "resize and fill":
                radio = r; break
        except: pass
    if not radio: return False
    try:
        checked = (radio.get_attribute("aria-checked") == "true") or radio.is_selected()
    except:
        checked = (radio.get_attribute("checked") is not None)
    if checked: return True
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", radio); time.sleep(0.03)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", radio),
        lambda: ActionChains(driver).move_to_element(radio).pause(0.02).click().perform(),
        lambda: radio.click(),
    ):
        try: fn(); time.sleep(0.06); break
        except: pass
    try:
        return (radio.get_attribute("aria-checked") == "true") or radio.is_selected()
    except:
        return (radio.get_attribute("checked") is not None)

def cn_unit_select_preproc_and_model(driver, unit_index: int, module_text: str | None, model_text: str | None) -> bool:
    if not _open_cn_unit_once(driver, unit_index): return False
    scope = _cn_unit_scope(driver, unit_index)
    if not scope: return False
    ok = False
    if module_text:
        ok |= select_listbox_scoped(driver, scope, "Preprocessor", module_text)
        if not ok: print(f"⚠️ Unit {unit_index} Preprocessor seçilemedi:", module_text)
    if model_text:
        ok2 = select_listbox_scoped(driver, scope, "Model", model_text)
        ok = ok or ok2
        if not ok2: print(f"⚠️ Unit {unit_index} Model seçilemedi:", model_text)
    return ok


# ------------------- Excel / Öğrenci Bilgisi -------------------
def _normalize_gender_token(raw: str | None) -> str:
    if not raw: return "boy"
    s = str(raw).strip().lower()
    if re.search(r"kız|kiz|female|woman|girl|kadın|kadin|f", s):
        return "girl"
    if re.search(r"erkek|male|man|boy|e", s):
        return "boy"
    return "boy"

def _find_gender_from_row(row: pd.Series) -> str | None:
    candidates = ["cinsiyet", "gender", "sex", "gender_", "sex_", "gender identity"]
    target_key = None
    for k in row.index:
        kl = str(k).strip().lower()
        if kl in candidates or "cins" in kl or "gender" in kl or kl == "sex":
            target_key = k; break
    if not target_key: return None
    return _normalize_gender_token(row.get(target_key))

def load_excel(excel_path: str | None) -> pd.DataFrame | None:
    if not excel_path or not os.path.exists(excel_path):
        return None
    try:
        return pd.read_excel(excel_path)
    except Exception:
        try:
            return pd.read_excel(excel_path, engine="openpyxl")
        except Exception:
            return None

def _pick_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    for c in df.columns:
        cl = str(c).strip().lower()
        if cl in candidates: return c
    for c in df.columns:
        cl = str(c).strip().lower()
        for key in candidates:
            if key in cl: return c
    return None

def student_info_from_excel(df: pd.DataFrame | None, face_path: str, settings: dict) -> tuple[str, str, str]:
    gender = "boy"
    cls = "ANA"
    name = "ÖĞRENCİ"
    if df is None:
        return cls, name, gender
    col_photo = settings.get("col_photo") or "@photo"
    col_class = settings.get("col_class") or None
    col_name  = settings.get("col_name")  or None
    if col_class is None:
        col_class = _pick_col(df, ["sınıf","sinif","class","branch","şube","sube","grup"])
    if col_name is None:
        col_name = _pick_col(df, ["ad soyad","adsoyad","isim","öğrenci adı","ogrenci adi","name","fullname","student","ogrenci","öğrenci"])
    row = None
    face_lower = os.path.abspath(face_path).lower()
    try:
        if col_photo in df.columns:
            row = df[df[col_photo].astype(str).str.lower() == face_lower].head(1)
            if row is not None and len(row)==0:
                row = None
    except Exception:
        row = None
    if row is None:
        fname = os.path.basename(face_lower)
        try:
            if col_photo in df.columns:
                row = df[df[col_photo].astype(str).str.lower().str.contains(re.escape(fname))].head(1)
                if row is not None and len(row)==0:
                    row = None
        except Exception:
            row = None
    if row is not None and len(row)>0:
        r = row.iloc[0]
        try:
            if col_class and col_class in df.columns and pd.notna(r.get(col_class)):
                cls = str(r.get(col_class)).strip()
        except: pass
        try:
            if col_name and col_name in df.columns and pd.notna(r.get(col_name)):
                name = str(r.get(col_name)).strip()
            else:
                c_ad   = _pick_col(df, ["ad","isim","first","adı","adi"])
                c_soy  = _pick_col(df, ["soyad","soyadı","soyadi","last","surname"])
                parts = []
                if c_ad and pd.notna(r.get(c_ad)): parts.append(str(r.get(c_ad)).strip())
                if c_soy and pd.notna(r.get(c_soy)): parts.append(str(r.get(c_soy)).strip())
                if parts: name = " ".join(parts)
        except: pass
        try:
            gtok = _find_gender_from_row(r)
            if gtok: gender = gtok
        except: pass
    return cls or "ANA", name or "ÖĞRENCİ", gender or "boy"

def _normalize_key(s: str) -> str:
    if s is None:
        return ""
    t = str(s).strip().lower()
    # Türkçe karakterler ve boşlukları normalize et
    tr_map = str.maketrans({
        "ç":"c","ğ":"g","ı":"i","ö":"o","ş":"s","ü":"u",
        "â":"a","î":"i","û":"u"
    })
    t = t.translate(tr_map)
    t = re.sub(r"\s+", "", t)  # boşluk sil
    t = t.replace("-", "").replace("_", "")
    return t

def _build_header_map(df: pd.DataFrame) -> dict:
    """
    DataFrame kolon adlarını normalize edip -> orijinal ada map eder.
    Aynı normalize anahtar iki kolona denk gelirse, ilkini korur.
    """
    m = {}
    for c in df.columns:
        k = _normalize_key(c)
        if k and k not in m:
            m[k] = c
    return m

def _find_row_by_face(df: pd.DataFrame, face_path: str, col_photo: str | None) -> pd.Series | None:
    """Foto sütunundan satır bul: önce tam yol, olmazsa dosya adıyla (case-insensitive)."""
    if df is None or not col_photo or col_photo not in df.columns:
        return None
    try:
        f_abs = os.path.abspath(face_path).lower()
        row = df[df[col_photo].astype(str).str.lower() == f_abs]
        if len(row) > 0:
            return row.iloc[0]
    except Exception:
        pass
    try:
        fname = os.path.basename(face_path).lower()
        row = df[df[col_photo].astype(str).str.lower().str.contains(re.escape(fname))]
        if len(row) > 0:
            return row.iloc[0]
    except Exception:
        pass
    return None


def resolve_placeholders(text: str, book: dict, face_path: str | None) -> str:
    """
    Prompt içindeki {SutunAdi} yer tutucularını Excel satırından çeker.
    Özel aliaslar:
      - {class}/{sınıf}/{sinif}  -> sınıf
      - {AdSoyad}                -> ad + soyad birleştirme (yoksa col_name)
      - {Cinsiyet}               -> 'boy'/'girl' normalizasyonu (mevcut mantık)
    Bulunamayanlar boş döner (settings.placeholder_default ile değiştirilebilir).
    """
    if not text:
        return text or ""

    settings = (book.get("settings") or {})
    df = load_excel(settings.get("excel_path") or settings.get("excel"))
    placeholder_default = settings.get("placeholder_default", "")  # istersen "-" vb.

    # Kullanıcı alias takımları
    CLASS_ALIASES = {_normalize_key("class"), _normalize_key("sınıf"), _normalize_key("sinif")}
    FULLNAME_KEY  = _normalize_key("AdSoyad")
    GENDER_KEY    = _normalize_key("Cinsiyet")

    # Satır ve başlık haritasını hazırla
    row = None
    header_map = {}
    col_photo = settings.get("col_photo") or "@photo"
    col_first = settings.get("col_first")
    col_last  = settings.get("col_last")
    col_name  = settings.get("col_name")
    col_class = settings.get("col_class")

    if df is not None:
        header_map = _build_header_map(df)
        if face_path:
            if col_photo and col_photo in df.columns:
                row = _find_row_by_face(df, face_path, col_photo)

    # Yardımcı: satırdan güvenli okuma
    def get_cell(colname: str) -> str:
        if row is None or df is None:
            return placeholder_default
        # 1) Tam sütun adı
        if colname in df.columns and pd.notna(row.get(colname)):
            return str(row.get(colname)).strip()
        # 2) Normalize ile eşleşen sütun
        norm = _normalize_key(colname)
        real = header_map.get(norm)
        if real and pd.notna(row.get(real)):
            return str(row.get(real)).strip()
        return placeholder_default

    # Sınıf değeri
    def get_class_value() -> str:
        val = ""
        # Öncelik: settings.col_class
        if col_class and df is not None and row is not None and col_class in df.columns and pd.notna(row.get(col_class)):
            val = str(row.get(col_class)).strip()
        # Yoksa öğrenci_info fallback (klasör ismi vs.):
        if not val:
            cls0, _, _ = student_info_from_excel(df, face_path or "", settings)
            val = cls0 or ""
        return val or placeholder_default

    # AdSoyad değeri
    def get_fullname_value() -> str:
        # 1) Tek kolon olarak tanımlanmışsa
        if col_name and df is not None and row is not None and col_name in df.columns and pd.notna(row.get(col_name)):
            return str(row.get(col_name)).strip()
        # 2) Ad + Soyad birleştir
        parts = []
        cand_first = col_first if (col_first and df is not None and col_first in df.columns) else None
        cand_last  = col_last  if (col_last  and df is not None and col_last  in df.columns) else None

        if not cand_first and df is not None:
            cand_first = header_map.get(_normalize_key("ad")) or header_map.get(_normalize_key("isim")) \
                       or header_map.get(_normalize_key("first")) or header_map.get(_normalize_key("adı")) \
                       or header_map.get(_normalize_key("adi"))
        if not cand_last and df is not None:
            cand_last  = header_map.get(_normalize_key("soyad")) or header_map.get(_normalize_key("soyadı")) \
                       or header_map.get(_normalize_key("soyadi")) or header_map.get(_normalize_key("last")) \
                       or header_map.get(_normalize_key("surname"))

        if row is not None and df is not None:
            if cand_first and pd.notna(row.get(cand_first)):
                parts.append(str(row.get(cand_first)).strip())
            if cand_last and pd.notna(row.get(cand_last)):
                parts.append(str(row.get(cand_last)).strip())

        if parts:
            return " ".join(parts)
        # 3) Hiçbiri yoksa student_info’dan gelen isim
        _, name0, _ = student_info_from_excel(df, face_path or "", settings)
        return (name0 or placeholder_default)

    # Cinsiyet değeri (mevcut normalize)
    def get_gender_value() -> str:
        if row is None or df is None:
            return "boy"
        tok = _find_gender_from_row(row)  # zaten 'boy'/'girl' döndürür
        return tok or "boy"

    # Tüm {…} yer tutucularını bul
    keys = set(re.findall(r"\{([^{}]+)\}", text or ""))  # içteki içerikleri al
    resolved = text

    for raw_key in keys:
        norm_key = _normalize_key(raw_key)

        # Özel anahtarlar
        if norm_key in CLASS_ALIASES:
            resolved = resolved.replace(f"{{{raw_key}}}", get_class_value())
            continue
        if norm_key == FULLNAME_KEY:
            resolved = resolved.replace(f"{{{raw_key}}}", get_fullname_value())
            continue
        if norm_key == GENDER_KEY:
            resolved = resolved.replace(f"{{{raw_key}}}", get_gender_value())
            continue

        # Genel durumda: Excel sütunu olarak ara
        val = get_cell(raw_key)
        resolved = resolved.replace(f"{{{raw_key}}}", val)

    return resolved



# ------------------- Styles + Generate + Çıktıyı Kaydet -------------------
def _find_styles_input(driver):
    labels = query_all(driver, "label")
    for lb in labels:
        try:
            txt = (lb.text or "").strip().lower()
            if "styles" in txt:
                inp = None
                try: inp = lb.find_element(By.CSS_SELECTOR, "input")
                except:
                    try:
                        inp = driver.execute_script("return arguments[0].querySelector('input');", lb)
                    except:
                        inp = None
                if inp:
                    return inp
        except:
            continue
    cands = query_all(driver, "input.border-none.subdued, input.border-none")
    for el in cands:
        try:
            if el.is_displayed() and (el.get_attribute("title") is not None):
                return el
        except:
            continue
    return None

def select_styles(driver, styles: list[str], clear_existing=True, pause=0.08):
    if not styles:
        return False
    inp = _find_styles_input(driver)
    if not inp:
        print("ℹ️ Styles input bulunamadı; atlanıyor.")
        return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", inp)
    except: pass
    time.sleep(pause)
    if clear_existing:
        try:
            inp.click(); time.sleep(pause)
            inp.send_keys(Keys.CONTROL, "a"); time.sleep(0.02)
            inp.send_keys(Keys.DELETE); time.sleep(pause)
        except: pass
    for st in styles:
        try:
            inp.click(); time.sleep(0.02)
            inp.send_keys(st)
            time.sleep(pause)
            inp.send_keys(Keys.ENTER)
            time.sleep(pause)
        except Exception as e:
            print(f"⚠️ Style seçilemedi: {st}  ({e})")
    return True

def click_generate(driver):
    btn = query_one(driver, "button#txt2img_generate") or query_one(driver, "#txt2img_generate")
    if not btn:
        for b in query_all(driver, "button"):
            try:
                t = (b.get_attribute("title") or "") + " " + (b.text or "")
                if "generate" in t.lower():
                    btn = b; break
            except:
                continue
    if not btn:
        print("⚠️ Generate butonu bulunamadı.")
        return False
    try:
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn); time.sleep(0.05)
    except: pass
    for fn in (
        lambda: driver.execute_script("arguments[0].click();", btn),
        lambda: ActionChains(driver).move_to_element(btn).pause(0.03).click().perform(),
        lambda: btn.click(),
    ):
        try:
            fn(); return True
        except: continue
    print("⚠️ Generate butonuna tıklanamadı.")
    return False

def _find_best_output_node(driver):
    return driver.execute_script("""
        function visible(el){
          if(!el) return false;
          const st=getComputedStyle(el);
          if(st.display==='none'||st.visibility==='hidden') return false;
          const r=el.getBoundingClientRect();
          return r.width>20 && r.height>20;
        }
        const sels=['#txt2img_gallery img','#txt2img_gallery canvas','div.gallery img','div.gallery canvas','img','canvas'];
        let best=null,bArea=0;
        for(const sel of sels){
            const nodes=document.querySelectorAll(sel);
            for(const el of nodes){
              if(!visible(el)) continue;
              const r=el.getBoundingClientRect();
              const area=r.width*r.height;
              if(area>bArea){ best=el; bArea=area; }
            }
        }
        return best || null;
    """)

def _node_to_dataurl(driver, node):
    if not node: return None
    kind = driver.execute_script("return arguments[0].tagName;", node)
    if kind and str(kind).upper()=="CANVAS":
        try:
            return driver.execute_script("return arguments[0].toDataURL('image/png');", node)
        except:
            pass
    try:
        src = driver.execute_script("return arguments[0].src || '';", node) or ""
    except:
        src = ""
    if src.startswith("data:image"):
        return src
    try:
        return driver.execute_script("""
            const img=arguments[0];
            const c=document.createElement('canvas');
            const w=img.naturalWidth||img.width, h=img.naturalHeight||img.height;
            if(!w||!h) return null;
            c.width=w; c.height=h;
            const ctx=c.getContext('2d');
            ctx.drawImage(img,0,0,w,h);
            return c.toDataURL('image/png');
        """, node)
    except:
        return None

def _sig_from_dataurl(dataurl: str) -> str:
    if not dataurl: return ""
    try:
        return hashlib.sha1(dataurl.encode("utf-8")).hexdigest()
    except:
        return dataurl[:120]

def snapshot_output_signature(driver) -> str:
    node = _find_best_output_node(driver)
    if not node:
        return ""
    dataurl = _node_to_dataurl(driver, node) or ""
    return _sig_from_dataurl(dataurl)

def _is_generate_running(driver) -> bool:
    """Buton, Stop, ilerleme çubuğu gibi sinyallerin herhangi biri aktifse True."""
    # 1) Generate butonu disable mı?
    btn = query_one(driver, "#txt2img_generate")
    try:
        if btn and (btn.get_attribute("disabled") is not None):
            return True
    except:
        pass

    # 2) Stop butonu görünüyor mu?
    for b in query_all(driver, "button"):
        try:
            if (b.text or "").strip().lower() == "stop":
                return True
        except:
            continue

    # 3) progressDiv görünür mü?
    if _is_progress_visible(driver):
        return True

    # 4) klasik progressbar’lar?
    try:
        prog = query_one(driver, 'div[role="progressbar"], div.progress-bar')
        if prog:
            st = (prog.get_attribute("style") or "").lower()
            if "visibility: hidden" not in st:
                return True
    except:
        pass
    return False


def wait_generation_cycle_and_save(driver, out_path: str, prev_signature: str, timeout=240) -> bool:
    """
    - progressDiv %100 (veya görünmez) olana kadar bekler
    - Öncelik: <img data-testid="detailed-image"> src içindeki yerel dosyayı kopyalar
    - Fallback: dataURL (img/canvas) alıp yazar
    """
    t_end = time.time() + timeout

    # 1) kısa süre içinde "çalışıyor" sinyali gör
    t0 = time.time() + 3.0
    while time.time() < t0:
        if _is_generate_running(driver):
            break
        time.sleep(0.1)

    # 2) progressDiv'i takip et
    saw_progress = False
    while time.time() < t_end:
        if _is_progress_visible(driver):
            saw_progress = True
            pct = _progress_percent(driver)
            if pct >= 100.0:
                time.sleep(0.6)
                break
        else:
            if not _is_generate_running(driver):
                break
        time.sleep(0.25)

    # 3) ÖNCE "detailed-image" img üzerinden doğrudan DOSYA KOPYALA
    t_find = time.time() + 20
    while time.time() < t_find:
        node = _find_final_image_node(driver)
        if node:
            try:
                src = node.get_attribute("src") or ""
            except:
                src = ""
            local_path = _src_to_local_path(src)
            if local_path and os.path.exists(local_path):
                try:
                    ensure_dir(Path(out_path).parent)
                    shutil.copyfile(local_path, out_path)
                    return True
                except Exception as e:
                    print("⚠️ Kopyalama hata:", e)
            # src yoksa/yerel yol çıkmadıysa hafif bekle ve tekrar dene
        time.sleep(0.3)

    # 4) Fallback: signature değişimine göre dataURL ile kaydet
    t_settle = time.time() + 30
    while time.time() < t_settle:
        node = _find_best_output_node(driver)
        if node:
            dataurl = _node_to_dataurl(driver, node)
            sig = _sig_from_dataurl(dataurl or "")
            if sig and sig != prev_signature and (dataurl or "").startswith("data:image"):
                try:
                    head, b64 = dataurl.split(",", 1)
                    raw = base64.b64decode(b64)
                    ensure_dir(Path(out_path).parent)
                    with open(out_path, "wb") as f:
                        f.write(raw)
                    return True
                except Exception as e:
                    print("⚠️ dataURL decode hata:", e)
        time.sleep(0.3)

    # 5) Son bir tetikleme
    if saw_progress:
        try:
            driver.execute_script("window.scrollBy(0, -120);"); time.sleep(0.15)
            driver.execute_script("window.scrollBy(0,  240);"); time.sleep(0.15)
        except: pass
        node = _find_final_image_node(driver) or _find_best_output_node(driver)
        if node:
            # tekrar img -> dosya kopya dene
            try:
                src = node.get_attribute("src") or ""
            except:
                src = ""
            local_path = _src_to_local_path(src)
            if local_path and os.path.exists(local_path):
                try:
                    ensure_dir(Path(out_path).parent)
                    shutil.copyfile(local_path, out_path)
                    return True
                except Exception as e:
                    print("⚠️ Kopyalama (fallback) hata:", e)
            # tekrar dataURL fallback dene
            dataurl = _node_to_dataurl(driver, node)
            if (dataurl or "").startswith("data:image"):
                try:
                    head, b64 = dataurl.split(",", 1)
                    raw = base64.b64decode(b64)
                    ensure_dir(Path(out_path).parent)
                    with open(out_path, "wb") as f:
                        f.write(raw)
                    return True
                except Exception as e:
                    print("⚠️ dataURL fallback decode hata:", e)

    print("⚠️ Yeni çıktı algılanamadı veya kaydedilemedi:", out_path)
    return False



# ------------------- keep-open (Enter) -------------------
def wait_for_enter_or_timeout(timeout_sec: float | None):
    try:
        if os.name == "nt":
            import msvcrt
            print("ℹ️ Pencereyi kapatmak için Enter’a basın...", flush=True)
            start = time.time()
            while True:
                if msvcrt.kbhit():
                    ch = msvcrt.getwch()
                    if ch in ("\r", "\n"):
                        break
                if timeout_sec is not None and (time.time() - start) >= timeout_sec:
                    print("⏱️ keep-open süresi doldu, kapanıyor...")
                    break
                time.sleep(0.08)
        else:
            input("ℹ️ Pencereyi kapatmak için Enter’a basın...")
    except Exception:
        try: input("ℹ️ Pencereyi kapatmak için Enter’a basın...")
        except: pass

# --- Control Weight + Control Mode yardımcıları (Integrated + Unit 1) ---

def _set_controlnet_weight_integrated(driver, value):
    """Integrated (Unit 0) alanında 'Control Weight' number input'u doldurur."""
    scope = _controlnet_scope(driver)
    if not scope: return False
    inp = _query_in_scope(driver, scope, 'input[aria-label="number input for Control Weight"][type="number"]') \
          or _query_in_scope(driver, scope, 'input[data-testid="number-input"][type="number"]')
    if not inp: return False
    try:
        driver.execute_script("arguments[0].value = arguments[1];", inp, str(value))
        fire(driver, inp)
        return True
    except:
        return False

def _select_controlnet_mode_integrated(driver, mode_text: str):
    """
    Integrated (Unit 0) alanında Control Mode radyo seçimi yapar.
    mode_text: 'Balanced' | 'My prompt is more important' | 'ControlNet is more important'
    """
    scope = _controlnet_scope(driver)
    if not scope or not mode_text: return False
    want = str(mode_text).strip().lower()
    radios = _query_all_in_scope(driver, scope, 'input[type="radio"]')
    for r in radios:
        try:
            val = (r.get_attribute("value") or "").strip().lower()
            if not val:
                lbl = driver.execute_script("return arguments[0].closest('label')", r)
                val = (lbl.text or "").strip().lower() if lbl else ""
            if want in val:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", r)
                try: driver.execute_script("arguments[0].click();", r)
                except:
                    try: ActionChains(driver).move_to_element(r).pause(0.02).click().perform()
                    except: r.click()
                time.sleep(0.06)
                return True
        except:
            continue
    return False

def _cn_unit_set_weight(driver, unit_index: int, value):
    """Unit {unit_index} alanında 'Control Weight' input'unu doldurur."""
    if not _open_cn_unit_once(driver, unit_index): return False
    scope = _cn_unit_scope(driver, unit_index)
    if not scope: return False
    inp = _query_in_scope(driver, scope, 'input[aria-label="number input for Control Weight"][type="number"]') \
          or _query_in_scope(driver, scope, 'input[data-testid="number-input"][type="number"]')
    if not inp: return False
    try:
        driver.execute_script("arguments[0].value = arguments[1];", inp, str(value))
        fire(driver, inp)
        return True
    except:
        return False

def _cn_unit_select_control_mode(driver, unit_index: int, mode_text: str):
    """Unit {unit_index} için Control Mode radyo seçimi."""
    if not _open_cn_unit_once(driver, unit_index): return False
    scope = _cn_unit_scope(driver, unit_index)
    if not scope or not mode_text: return False
    want = str(mode_text).strip().lower()
    radios = _query_all_in_scope(driver, scope, 'input[type="radio"]')
    for r in radios:
        try:
            val = (r.get_attribute("value") or "").strip().lower()
            if not val:
                lbl = driver.execute_script("return arguments[0].closest('label')", r)
                val = (lbl.text or "").strip().lower() if lbl else ""
            if want in val:
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", r)
                try: driver.execute_script("arguments[0].click();", r)
                except:
                    try: ActionChains(driver).move_to_element(r).pause(0.02).click().perform()
                    except: r.click()
                time.sleep(0.06)
                return True
        except:
            continue
    return False

def _mode_int_to_text(v: int) -> str:
    return {
        0: "Balanced",
        1: "My prompt is more important",
        2: "ControlNet is more important",
    }.get(int(v or 0), "Balanced")




# ------------------- Tek Sayfa Workflow (driver paylaşılabilir) -------------------
def fill_prompts_and_basic_params(
    forge_url: str,
    prompt: str,
    neg_prompt: str,
    width=None, height=None, steps=None, cfg_scale=None, seed=None,
    sampler_name=None,
    use_reactor=False,
    use_controlnet=False,
    face_path: str | None = None,
    faces_dir_fallback: str | None = None,
    headless: bool = False,
    ensure_txt2img=True,
    initial_delay_sec: float = 1.8,
    # Unit 0
    cn_module_text: str | None = None,
    cn_model_text: str | None = None,
    # Unit 1
    cn1_module_text: str | None = None,
    cn1_model_text: str | None = None,
    pose_path: str | None = None,
    # Book for placeholders
    book: dict | None = None,
    styles: list[str] | None = None,
    # === Control Weight + Control Mode ===
    cn0_weight: float | None = 0.5,
    cn1_weight: float | None = 0.5,
    cn0_control_mode: str | None = None,   # 'Balanced' | 'My prompt is more important' | 'ControlNet is more important'
    cn1_control_mode: str | None = None,
    # === Resize modları (0=Just, 1=Crop, 2=Fill)
    cn0_resize_mode: int | None = None,
    cn1_resize_mode: int | None = None,
    # KAYDETME
    save_image_to: str | None = None,
    save_timeout_sec: int = 180,
    # driver yönetimi
    driver=None,
    manage_driver: bool = True,
):
    own_driver = False
    if driver is None:
        driver  = new_driver(headless=headless)
        own_driver = True

    wait = WebDriverWait(driver, 30)
    try:
        driver.get(forge_url)
        time.sleep(initial_delay_sec)
        to_fullscreen(driver)

        wait.until(lambda d: _get_app(d) or query_one(d, 'textarea[placeholder^="Prompt"]'))
        if ensure_txt2img:
            maybe_switch_to_txt2img(driver); time.sleep(0.15)

        wait.until(lambda d: find_prompt_textareas(d)[0] is not None)
        p_el, n_el = find_prompt_textareas(driver)
        if not p_el:  raise RuntimeError("Prompt textarea bulunamadı.")
        if not n_el:  raise RuntimeError("Negative prompt textarea bulunamadı.")

        prompt = resolve_placeholders(prompt, book or {}, face_path)
        set_text(driver, p_el, prompt); set_text(driver, n_el, neg_prompt)

        w_el = query_one(driver, 'input[aria-label^="number input for Width"], input[aria-label="Width"], input#txt2img_width, input#width')
        h_el = query_one(driver, 'input[aria-label^="number input for Height"], input[aria-label="Height"], input#txt2img_height, input#height')
        if w_el and width is not None:   set_number(driver, w_el, width)
        if h_el and height is not None:  set_number(driver, h_el, height)

        st_el  = query_one(driver, 'input[aria-label^="number input for Sampling steps"], input[aria-label="Steps"], input#txt2img_steps, input#steps')
        if st_el and steps is not None:  set_number(driver, st_el, steps)

        cfg_el = query_one(driver, 'input[aria-label^="number input for CFG Scale"], input[aria-label="CFG Scale"], input[aria-label="CFG scale"], input#txt2img_cfg_scale, input#cfg_scale')
        if cfg_el and cfg_scale is not None: set_float(driver, cfg_el, cfg_scale)

        sd_el  = query_one(driver, 'input[aria-label^="number input for Seed"], input[aria-label="Seed"], input#seed')
        if sd_el and seed is not None and str(seed).strip() != "" and int(seed) >= 0:
            set_number(driver, sd_el, seed)

        ok_sampler = False
        samp_sel = query_one(driver, 'select[aria-label="Sampler"], select#txt2img_sampling, select#sampler, select#sampling')
        if samp_sel and sampler_name:
            ok_sampler = set_select_by_text_or_value(driver, samp_sel, sampler_name)
        if not ok_sampler and sampler_name:
            ok_sampler = open_dropdown_and_pick(driver, "Sampling method", sampler_name)
            if not ok_sampler:
                print("⚠️ Sampling method seçilemedi:", sampler_name)

        # Styles
        try:
            if styles and len(styles) > 0:
                select_styles(driver, styles, clear_existing=True)
        except Exception as e:
            print("⚠️ Styles seçimi hata:", e)

        # REActor
        if use_reactor:
            upload_path = face_path or None
            if upload_path and os.path.exists(upload_path):
                ok = upload_face_in_reactor_panel(driver, upload_path)
                if not ok:
                    btn = _find_reactor_button(driver)
                    cont = _reactor_container_from_button(driver, btn) if btn else None
                    if not _reactor_upload_confirmed(driver, cont, None):
                        print("⚠️ REActor file upload başarısız.")
            else:
                print("ℹ️ REActor için yüz görseli bulunamadı.")

        # ------------------ ControlNet ------------------
        if use_controlnet:
            if not ensure_controlnet_checkbox_on(driver):
                print("⚠️ ControlNet kutusu açılamadı/işaretlenemedi.")
            else:
                # ---- Unit 0 (Integrated) ----
                # seçilen preprocessor’a göre radyo tercihi
                is_instant0  = _is_instant_module(cn_module_text or "")
                is_openpose0 = _is_openpose_module(cn_module_text or "")

                if is_instant0:
                    try: select_instant_id_radio(driver)
                    except Exception: print("⚠️ Instant-ID radio seçimi denemesi başarısız.")
                elif is_openpose0:
                    try: select_openpose_radio(driver)
                    except Exception: print("⚠️ OpenPose radio seçimi denemesi başarısız.")
                # değilse radio zorlamıyoruz

                # Unit0: preprocessor & model
                select_controlnet_preproc_and_model(driver, cn_module_text, cn_model_text)

                # Unit0: image upload
                upload_path = face_path or faces_dir_fallback
                if upload_path and os.path.exists(upload_path):
                    if not upload_student_image_to_controlnet(driver, upload_path):
                        print("⚠️ ControlNet iç resim yükleme başarısız.")

                # Unit0: weight + mode
                if cn0_weight is not None:
                    try: _set_controlnet_weight_integrated(driver, cn0_weight)
                    except Exception as e: print("⚠️ Unit0 weight set hata:", e)
                if cn0_control_mode:
                    try: _select_controlnet_mode_integrated(driver, cn0_control_mode)
                    except Exception as e: print("⚠️ Unit0 mode set hata:", e)

                # Unit0: resize mode (varsa uygula)
                if cn0_resize_mode is not None:
                    ok_r0 = _set_resize_mode_integrated(driver, int(cn0_resize_mode))
                    if not ok_r0:
                        print("⚠️ Unit 0 resize mode seçilemedi.")

            # ---- Unit 1 (ek panel) ----
            UNIT = 1
            if ensure_cn_unit_checkbox_on(driver, UNIT):
                # Preproc’a göre radio (opsiyonel)
                if _is_instant_module(cn1_module_text or ""):
                    cn_unit_select_instant_id(driver=driver, unit_index=UNIT)
                elif _is_openpose_module(cn1_module_text or ""):
                    try:
                        cont = _cn_unit_get_container(driver, UNIT)
                        _click_radio_by_value_in(cont or driver, "OpenPose")
                    except Exception:
                        pass

                # Preproc & model
                cn_unit_select_preproc_and_model(driver, UNIT, cn1_module_text, cn1_model_text)

                # Image upload (pose varsa)
                if pose_path and os.path.exists(pose_path):
                    cn_unit_upload_image(driver, UNIT, pose_path)

                # Unit1: resize mode
                if cn1_resize_mode is not None:
                    ok_r1 = _cn_unit_select_resize_mode(driver, UNIT, int(cn1_resize_mode))
                    if not ok_r1:
                        print("⚠️ Unit 1 resize mode seçilemedi.")
                else:
                    # geri uyumluluk: eskisi gibi Resize and Fill
                    cn_unit_select_resize_and_fill(driver, UNIT)

                # Unit1: weight + mode
                if cn1_weight is not None:
                    try: _cn_unit_set_weight(driver, UNIT, cn1_weight)
                    except Exception as e: print("⚠️ Unit1 weight set hata:", e)
                if cn1_control_mode:
                    try: _cn_unit_select_control_mode(driver, UNIT, cn1_control_mode)
                    except Exception as e: print("⚠️ Unit1 mode set hata:", e)

        # ------------------ Çalıştır & Kaydet ------------------
        prev_sig = snapshot_output_signature(driver)
        clicked = click_generate(driver)
        if not clicked:
            print("⚠️ Generate tıklanamadı.")
        if save_image_to:
            ok_save = wait_generation_cycle_and_save(driver, save_image_to, prev_sig, timeout=save_timeout_sec)
            if ok_save:
                print(f"💾 Kaydedildi: {save_image_to}")
            else:
                print(f"⚠️ Kaydedilemedi: {save_image_to}")

        print("✅ Sayfa dolduruldu ve Generate çalıştı.")
        return True

    finally:
        if own_driver and manage_driver:
            try: driver.quit()
            except: pass


# ------------------- Otomasyon / Batch -------------------
def slugify_for_path(s: str) -> str:
    s = s or ""
    s = s.replace(":", " ").replace("/", " ").replace("\\"," ").replace("|"," ").replace("*"," ")
    s = re.sub(r'\s+', ' ', s).strip()
    return s

def run_batch(book: dict, forge_url: str, headless=False, initial_delay=1.8):
    """
    Excel/CSV varsa çocukları satır sırasına göre, yoksa faces_dir hiyerarşisine göre sırayla işler.
    KALDIĞI YERDEN DEVAM:
      - output_root/<Sınıf>/<Ad Soyad>/sayfa{N}.png mevcutsa o sayfa atlanır
      - Tüm sayfaları mevcut olan çocuk atlanır

    NOT:
    - Excel kipinde 'out' SÜTUNU KULLANILMAZ / OLUŞTURULMAZ.
    - '@sayfa1..@sayfaN' başlıkları bir kez, ilk tamamen boş sütun bloğundan başlayacak şekilde yerleştirilir
      ve tüm satırlar bu sabit konumları kullanır.
    """
    settings    = (book.get("settings") or {})
    faces_dir   = settings.get("faces_dir")
    output_root = (settings.get("output_root") or str(ROOT_DIR / "out")).strip()
    excel_path  = (settings.get("excel_path") or settings.get("excel") or "").strip()
    data_source = (settings.get("data_source") or "excel").strip().lower()

    pages = sorted(book.get("pages", []), key=lambda x: x.get("index", 0))
    total_pages = len(pages)

    def page_out_path(base_dir: Path, page_index: int) -> Path:
        base_dir.mkdir(parents=True, exist_ok=True)
        return base_dir / f"sayfa{int(page_index)}.png"

    # ---------------- Excel Writer (yalın, 'out' yok) ----------------
    class _ExcelOutWriter:
        """
        - Başlık satırında '@sayfa1..@sayfaN' blokunu sadece 1 kez yerleştirir.
        - 'out' kolonuna dokunmaz (varsa bile yazmaz, yoksa yaratmaz).
        """
        def __init__(self, path: str):
            import csv
            self.path = path
            self.ext = os.path.splitext(path)[1].lower()
            self.mode = "xlsx" if self.ext in (".xlsx", ".xlsm") else "csv"
            self.valid = False
            self.pages_start_col = None         # 1-based
            self.pages_count = 0
            if not os.path.exists(path):
                print("⚠️ ExcelOutWriter: dosya yok, yazma atlanacak:", path)
                return
            try:
                if self.mode == "xlsx":
                    from openpyxl import load_workbook
                    self.wb = load_workbook(path)
                    self.ws = self.wb.active
                    header_cells = list(self.ws.iter_rows(min_row=1, max_row=1))[0]
                    self.header = [("" if c.value is None else str(c.value)) for c in header_cells]
                else:
                    with open(path, newline="", encoding="utf-8-sig") as f:
                        rows = list(csv.reader(f))
                    if not rows:
                        rows = [[]]
                    self.header = rows[0]
                    self.rows = rows[1:]
                self.valid = True
            except Exception as e:
                print("⚠️ ExcelOutWriter: açılamadı:", e)
                self.valid = False

        # ---- header yardımcıları ----
        def _first_empty_block_start(self, need_cols: int) -> int:
            """1-based: başlık satırındaki ilk ardışık boş blok başlangıcını bulur; yoksa sona ekler."""
            if need_cols <= 0:
                return len(self.header) + 1
            n = len(self.header)
            norm = [((c or "").strip()) for c in self.header]
            run = 0
            start = 1
            for idx in range(1, n + 1):
                if norm[idx-1] == "":
                    if run == 0:
                        start = idx
                    run += 1
                    if run >= need_cols:
                        return start
                else:
                    run = 0
            return n + 1

        def _ensure_header_len(self, new_len: int):
            if self.mode == "xlsx":
                # openpyxl tarafında header listemizi genişletelim (sheet hücreleri yazıldıkça fiilen oluşur)
                if new_len > len(self.header):
                    self.header += [""] * (new_len - len(self.header))
            else:
                if new_len > len(self.header):
                    self.header += [""] * (new_len - len(self.header))

        def init_pages_block_once(self, pages_count: int):
            """Başlıkları bir defa yerleştirir; yeniden çağrılırsa mevcut bloğu aynen kullanır."""
            if not self.valid:
                return
            if self.pages_start_col and self.pages_count == pages_count:
                return  # zaten hazır

            # Eğer header içinde zaten tam bir '@sayfa1..N' bloğu varsa onu kullan
            # (tek tek kontrol; toleranslı)
            def _scan_existing():
                labels = [f"@sayfa{i}" for i in range(1, pages_count+1)]
                norm = [((c or "").strip().lstrip("'")) for c in self.header]  # baştaki tek tırnak varsa temizle
                for start in range(1, len(norm) - pages_count + 2):
                    ok = True
                    for j, lab in enumerate(labels, start=0):
                        if (start-1+j) >= len(norm) or norm[start-1+j].lower() != lab.lower():
                            ok = False
                            break
                    if ok:
                        return start
                return None

            existing_start = _scan_existing()
            if existing_start:
                self.pages_start_col = existing_start
                self.pages_count = pages_count
                return

            # Yoksa: ilk tamamen boş bloktan başlat
            start_col = self._first_empty_block_start(pages_count)
            end_col = start_col + pages_count - 1
            self._ensure_header_len(end_col)

            titles = [f"@sayfa{i}" for i in range(1, pages_count+1)]
            if self.mode == "xlsx":
                for j, title in enumerate(titles, start=start_col):
                    # başta tek tırnak YOK (Excel metin kabul ediyor zaten)
                    self.ws.cell(row=1, column=j, value=title)
                    # header cache
                    self.header[j-1] = title
            else:
                for j, title in enumerate(titles, start=start_col):
                    self.header[j-1] = title

            self.pages_start_col = start_col
            self.pages_count = pages_count

        # ---- satır yazımı ----
        def set_pages_for_row(self, row_index_2based: int, page_paths: list[str]):
            if not self.valid or not self.pages_start_col or row_index_2based < 2:
                return
            # Listeyi tam N uzunluğa normalize et (eksik sayfalar "")
            vals = list(page_paths or [])
            if len(vals) < self.pages_count:
                vals = vals + [""] * (self.pages_count - len(vals))
            else:
                vals = vals[:self.pages_count]

            if self.mode == "xlsx":
                for k, p in enumerate(vals, start=0):
                    self.ws.cell(row=row_index_2based, column=self.pages_start_col + k, value=p)
            else:
                i = row_index_2based - 2
                while i >= len(self.rows):
                    self.rows.append([])
                # satırı header uzunluğuna kadar büyüt
                if len(self.rows[i]) < len(self.header):
                    self.rows[i] += [""] * (len(self.header) - len(self.rows[i]))
                for k, p in enumerate(vals, start=0):
                    idx = self.pages_start_col - 1 + k
                    if idx >= len(self.rows[i]):
                        self.rows[i] += [""] * (idx - len(self.rows[i]) + 1)
                    self.rows[i][idx] = p

        def save(self):
            if not self.valid:
                return
            try:
                if self.mode == "xlsx":
                    self.wb.save(self.path)
                else:
                    import csv
                    with open(self.path, "w", newline="", encoding="utf-8-sig") as f:
                        w = csv.writer(f)
                        w.writerow(self.header)
                        for r in self.rows:
                            if len(r) < len(self.header):
                                r = r + [""] * (len(self.header) - len(r))
                            w.writerow(r)
            except Exception as e:
                print("⚠️ ExcelOutWriter.save hata:", e)

    # ---------------- çocukları hazırla ----------------
    children = []
    writer = None
    df = None
    used_excel = False

    if data_source == "excel" and excel_path and os.path.exists(excel_path):
        try:
            import pandas as pd
            ext = os.path.splitext(excel_path)[1].lower()
            if ext == ".csv":
                for enc in ("utf-8-sig", "utf-8", "cp1254", "latin-1"):
                    try:
                        df = pd.read_csv(excel_path, encoding=enc); break
                    except Exception:
                        df = None
            else:
                try:
                    df = pd.read_excel(excel_path)
                except Exception:
                    try:
                        df = pd.read_excel(excel_path, engine="openpyxl")
                    except Exception:
                        df = None
        except Exception as e:
            print("⚠️ Excel/CSV okuma hatası, klasör kipine düşülecek:", e)
            df = None

        if df is not None and len(df) > 0:
            used_excel = True
            writer = _ExcelOutWriter(excel_path)
            if writer and writer.valid:
                # BAŞLIKLARI BİR KEZ sabitle
                writer.init_pages_block_once(total_pages)

            def _pick_ci(df, cand):
                if not cand: return None
                c = cand.strip().lower()
                for col in df.columns:
                    if str(col).strip().lower() == c:
                        return col
                return None

            col_photo = _pick_ci(df, settings.get("col_photo") or "@photo") or _pick_ci(df, "@photo") or _pick_ci(df, "photo")
            col_first = _pick_ci(df, settings.get("col_first") or "")
            col_last  = _pick_ci(df, settings.get("col_last")  or "")
            col_class = _pick_ci(df, settings.get("col_class") or "")

            def _get(row, col, default=""):
                try:
                    import pandas as _pd
                    if not col: return default
                    v = row.get(col)
                    return "" if _pd.isna(v) else str(v).strip()
                except Exception:
                    return default

            def _join_face(rel_or_abs: str) -> str:
                if not rel_or_abs: return ""
                s = str(rel_or_abs).replace("/", os.sep).replace("\\", os.sep).strip()
                if os.path.isabs(s): return os.path.normpath(s)
                if faces_dir:       return os.path.normpath(os.path.join(faces_dir, s))
                return os.path.normpath(s)

            for i, row in enumerate(df.to_dict(orient="records"), start=2):
                rel = _get(row, col_photo, "") if col_photo else ""
                if not rel:
                    continue
                face_abs = _join_face(rel)
                if not os.path.exists(face_abs):
                    print(f"⚠️ Yüz dosyası yok (Excel satır {i}): {face_abs}")
                    continue
                first = _get(row, col_first, "")
                last  = _get(row, col_last, "")
                cls   = _get(row, col_class, "")
                name  = " ".join([x for x in [first, last] if x]).strip() or os.path.splitext(os.path.basename(face_abs))[0]
                children.append({"face": face_abs, "class": cls or "ANA", "name": name, "row_index": i})

    if not used_excel:
        faces = list_faces_in_dir(faces_dir)
        if not faces:
            print("⚠️ faces_dir içinde işlenecek görsel bulunamadı:", faces_dir)
            return
        for fp in faces:
            cls  = Path(fp).parent.name or "ANA"
            name = os.path.splitext(Path(fp).name)[0]
            children.append({"face": fp, "class": cls, "name": name, "row_index": None})

    drv = new_driver(headless=headless)
    to_fullscreen(drv)
    print(f"🧒 Öğrenci sayısı: {len(children)}  |  Sayfa adedi: {len(pages)}  |  Kaynak: {'Excel' if used_excel else 'Klasör'}")

    # Excel DataFrame (placeholder lookup için)
    df_for_lookup = df if used_excel else load_excel(excel_path)

    for idx_child, ch in enumerate(children, start=1):
        face_path = ch["face"]
        cls0, name0, _ = student_info_from_excel(df_for_lookup, face_path, settings)
        cls  = slugify_for_path(ch.get("class") or cls0 or "ANA")
        name = slugify_for_path(ch.get("name")  or name0 or "ÖĞRENCİ")
        child_base = Path(output_root) / cls / name

        # Bitmiş sayfaları saptama
        done_map = {}
        for pg in pages:
            pidx = int(pg.get("index", 0) or 0)
            done_map[pidx] = page_out_path(child_base, pidx).exists()
        all_done = all(done_map.values()) if done_map else False
        if all_done:
            print(f"\n=== [{idx_child}/{len(children)}] {cls} / {name} → TÜM SAYFALAR VAR, ATLANIYOR ===")
            # 'out' sütununa kesinlikle yazma (legacy kapalı)
            # Ancak '@sayfa*' başlıkları zaten sabit; satırda mevcut yolları güncelle:
            if used_excel and writer and ch.get("row_index"):
                try:
                    existing = []
                    for pg in pages:
                        pidx = int(pg.get("index", 0) or 0)
                        pth = page_out_path(child_base, pidx)
                        if pth.exists():
                            existing.append(str(pth))
                    writer.set_pages_for_row(ch["row_index"], existing)
                    writer.save()
                except Exception as e:
                    print("⚠️ Excel '@sayfa*' (skip) yazılamadı:", e)
            continue

        print(f"\n=== [{idx_child}/{len(children)}] {cls} / {name} ===")

        # Forge her çocukta tazelensin
        drv.get(forge_url)
        time.sleep(initial_delay)
        maybe_switch_to_txt2img(drv); time.sleep(0.15)

        # Satır için toplanacak yollar
        page_paths_for_row = []
        # Önce mevcutları sıraya koyalım (1..N)
        existing_map = {}
        for pg in pages:
            pidx = int(pg.get("index", 0) or 0)
            out_p = page_out_path(child_base, pidx)
            if out_p.exists():
                existing_map[pidx] = str(out_p)

        for pg in pages:
            pidx = int(pg.get("index", 0) or 0)
            out_p = page_out_path(child_base, pidx)
            if out_p.exists():
                print(f" -> Sayfa #{pidx} ATLA (mevcut): {out_p}")
                continue

            print(f" -> Sayfa #{pidx}  (çıktı: {out_p})")

            prompt      = pg.get("prompt","")
            neg         = pg.get("negative_prompt","")
            width       = pg.get("width", None)
            height      = pg.get("height", None)
            steps       = pg.get("sampling_steps", None)
            cfg_scale   = pg.get("cfg_scale", None)
            seed        = pg.get("seed", None)
            sampler     = (pg.get("sampling_method") or "").strip()
            styles_list = pg.get("styles") or []

            use_reactor   = bool(pg.get("use_reactor"))
            use_control   = bool(pg.get("use_controlnet")) or bool(settings.get("ui_use_controlnet"))
            cn_module_txt = pg.get("cn0_module") or ""
            cn_model_txt  = pg.get("cn0_model")  or ""
            cn1_mod_txt   = pg.get("cn1_module") or ""
            cn1_model_txt = pg.get("cn1_model")  or ""
            pose_path     = pg.get("pose_path")  or ""

            fill_prompts_and_basic_params(
                forge_url=forge_url,
                prompt=prompt,
                neg_prompt=neg,
                width=width,
                height=height,
                steps=steps,
                cfg_scale=cfg_scale,
                seed=seed,
                sampler_name=sampler,
                use_reactor=use_reactor,
                use_controlnet=use_control,
                face_path=face_path,
                faces_dir_fallback=faces_dir,
                headless=headless,
                ensure_txt2img=True,
                initial_delay_sec=0.2,
                cn_module_text=cn_module_txt,
                cn_model_text=cn_model_txt,
                cn1_module_text=cn1_mod_txt,
                cn1_model_text=cn1_model_txt,
                pose_path=pose_path,
                book=book,
                styles=styles_list,
                cn0_weight=float(pg.get("cn0_weight", 0.5)),
                cn1_weight=float(pg.get("cn1_weight", 0.5)),
                cn0_control_mode=_mode_int_to_text(pg.get("cn0_mode", 0)),
                cn1_control_mode=_mode_int_to_text(pg.get("cn1_mode", 0)),
                cn0_resize_mode=int(pg.get("cn0_resize", 1)),
                cn1_resize_mode=int(pg.get("cn1_resize", 2)),
                save_image_to=str(out_p),
                save_timeout_sec=300,
                driver=drv,
                manage_driver=False,
            )

            # yeni dosya oluştuysa kayda geç
            if out_p.exists():
                existing_map[pidx] = str(out_p)

        # Satır yazımı: 1..N sıraya göre liste oluştur
        for pg in pages:
            pidx = int(pg.get("index", 0) or 0)
            page_paths_for_row.append(existing_map.get(pidx, ""))

        if used_excel and writer and ch.get("row_index"):
            try:
                writer.set_pages_for_row(ch["row_index"], page_paths_for_row)
                writer.save()
                if any(page_paths_for_row):
                    print(f"📝 Excel '@sayfa*' yazıldı (satır {ch['row_index']}).")
            except Exception as e:
                print("⚠️ Excel '@sayfa*' yazılamadı:", e)

    try:
        if writer:
            writer.save()
    except Exception:
        pass

    try:
        drv.quit()
    except Exception:
        pass




def _progress_div(driver):
    try:
        return query_one(driver, "div.progressDiv")
    except:
        return None

def _progress_percent(driver):
    """progressDiv içindeki .progress width % değerini döndürür; yoksa -1."""
    div = _progress_div(driver)
    if not div:
        return -1
    try:
        bar = driver.execute_script("return arguments[0].querySelector('div.progress');", div)
        if not bar:
            return -1
        style = (bar.get_attribute("style") or "")
        m = re.search(r"width:\s*([0-9.]+)%", style)
        if not m:
            return -1
        return float(m.group(1))
    except:
        return -1

def _is_progress_visible(driver):
    div = _progress_div(driver)
    if not div:
        return False
    try:
        st = (div.get_attribute("style") or "").lower()
        # display: none değilse görünür kabul
        return "display: none" not in st
    except:
        return False


import shutil
import urllib.parse

def _find_final_image_node(driver):
    """
    Forge'un detay çıktısı: <img data-testid="detailed-image">.
    Görünür en büyük 'detailed-image' <img>'i döndürür.
    """
    cands = query_all(driver, 'img[data-testid="detailed-image"]')
    best, area = None, 0
    for img in cands:
        try:
            st = (img.get_attribute("style") or "").lower()
            if "display: none" in st or "visibility: hidden" in st:
                continue
            r = driver.execute_script("return arguments[0].getBoundingClientRect()", img)
            a = max(0.0, float(r.get("width", 0))) * max(0.0, float(r.get("height", 0)))
            if a > area:
                best, area = img, a
        except:
            continue
    return best

def _src_to_local_path(src: str) -> str | None:
    """
    http://127.0.0.1:7861/file=C:\...\foo.png
    veya
    http://127.0.0.1:7861/file=C%3A%5C...\foo.png
    gibi URL'lerden yerel dosya yolunu çıkarır.
    """
    if not src:
        return None
    # parçala
    try:
        # query değil, doğrudan path’in içinde 'file=' var; onu ara
        idx = src.find("file=")
        if idx >= 0:
            raw = src[idx+5:]
        else:
            raw = src
        # '#' vb. olasılıklarını temizle
        raw = raw.split("#", 1)[0]
        # URL decode
        raw = urllib.parse.unquote(raw)
        # Bazı durumlarda forward slash gelebilir; Windows’a çevir
        if re.match(r"^[A-Za-z]:/", raw):
            raw = raw.replace("/", "\\")
        # Emin olmak için dosya var mı?
        return raw if os.path.exists(raw) else None
    except:
        return None




# ------------------- CLI -------------------
def main():
    ap = argparse.ArgumentParser(description="WebUI Forge otomasyon: faces_dir -> tüm sayfalar; REActor/ControlNet + Styles + Generate + Save")
    ap.add_argument("--book-id", required=True, help="data/books/<book_id>.json")
    g = ap.add_mutually_exclusive_group()
    g.add_argument("--page-index", type=int, help="Sadece tek sayfa çalış (1-based)")
    g.add_argument("--page-id", help="Sadece tek sayfa çalış (id)")
    ap.add_argument("--forge-url", default="http://127.0.0.1:7861/", help="Forge URL")
    ap.add_argument("--headless", action="store_true", help="Headless tarayıcı")
    ap.add_argument("--keep-open", action="store_true", help="(Tek sayfa modunda) İş bittiğinde tarayıcı açık kalsın (Enter ile kapanır)")
    ap.add_argument("--keep-open-timeout", type=float, default=None, help="Enter beklerken otomatik kapanma süresi (sn)")
    ap.add_argument("--face-path", default=None, help="Tek sayfa modunda REActor/ControlNet için yüz/öğrenci görseli")
    ap.add_argument("--initial-delay", type=float, default=1.8, help="İlk yükleme beklemesi (sn)")
    ap.add_argument("--batch", action="store_true", help="Otomasyon: faces_dir altındaki TÜM öğrenciler ve TÜM sayfalar")
    # ... mevcut argümanların altına ekle ...
    ap.add_argument("--children-json", type=str, default="",
                    help="Excel sırası manifest JSON (app.py tarafından üretilir)")

    args = ap.parse_args()

    book = load_book(args.book_id)
    settings = (book.get("settings") or {})

    if args.batch:
        run_batch(book, forge_url=args.forge_url, headless=args.headless, initial_delay=args.initial_delay)
        return

    page = pick_page(book, args.page_index, args.page_id)

    prompt      = page.get("prompt","")
    neg         = page.get("negative_prompt","")
    width       = page.get("width", None)
    height      = page.get("height", None)
    steps       = page.get("sampling_steps", None)
    cfg_scale   = page.get("cfg_scale", None)
    seed        = page.get("seed", None)
    sampler     = (page.get("sampling_method") or "").strip()
    styles_list = page.get("styles") or []

    use_reactor = bool(page.get("use_reactor"))
    use_control = bool(page.get("use_controlnet")) or bool(settings.get("ui_use_controlnet"))
    faces_dir   = settings.get("faces_dir")

    cn_module_text = page.get("cn0_module") or ""
    cn_model_text  = page.get("cn0_model")  or ""
    cn1_module_text = page.get("cn1_module") or ""
    cn1_model_text  = page.get("cn1_model")  or ""
    pose_path       = page.get("pose_path")  or ""

    output_root = settings.get("output_root") or str(ROOT_DIR / "out")
    df = load_excel(settings.get("excel_path") or settings.get("excel"))
    cls, name, _ = student_info_from_excel(df, args.face_path or "", settings)
    out_path = (
            Path(output_root)
            / slugify_for_path(cls)
            / slugify_for_path(name)
            / f"sayfa{int(page.get('index', 0) or 0)}.png"
    )

    print(f"Kitap: {book.get('name','-')}  |  Sayfa: #{page.get('index')}  ({page.get('id')})")
    print(f"Sampler: {sampler or '-'}  |  REActor: {'Açık' if use_reactor else 'Kapalı'}  |  ControlNet: {'Açık' if use_control else 'Kapalı'}")

    fill_prompts_and_basic_params(
        forge_url=args.forge_url,
        prompt=prompt,
        neg_prompt=neg,
        width=width,
        height=height,
        steps=steps,
        cfg_scale=cfg_scale,
        seed=seed,
        sampler_name=sampler,
        use_reactor=use_reactor,
        use_controlnet=use_control,
        face_path=args.face_path,
        faces_dir_fallback=faces_dir,
        headless=args.headless,
        ensure_txt2img=True,
        initial_delay_sec=args.initial_delay,
        cn_module_text=cn_module_text,
        cn_model_text=cn_model_text,
        cn1_module_text=cn1_module_text,
        cn1_model_text=cn1_model_text,
        pose_path=pose_path,
        book=book,
        styles=styles_list,
        # --- Control Weights & Modes ---
        cn0_weight=float(page.get("cn0_weight", 0.5)),
        cn1_weight=float(page.get("cn1_weight", 0.5)),
        cn0_control_mode=_mode_int_to_text(page.get("cn0_mode", 0)),
        cn1_control_mode=_mode_int_to_text(page.get("cn1_mode", 0)),
        # --- YENİ: Resize modları (0=Just, 1=Crop, 2=Fill) ---
        cn0_resize_mode=int(page.get("cn0_resize", 1)),
        cn1_resize_mode=int(page.get("cn1_resize", 2)),
        # --- Kayıt ---
        save_image_to=str(out_path),
        save_timeout_sec=300,
        driver=None,
        manage_driver=True,
    )

    if args.keep_open and not args.headless:
        wait_for_enter_or_timeout(args.keep_open_timeout)

if __name__ == "__main__":
    main()
